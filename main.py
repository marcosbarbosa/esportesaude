# ==============================================================================
# 📄 Arquivo: main.py novo modulo
# 🏷️ VERSÃO: 14.5 (PRO Elite - Fix CSS Dark Mode e Integração Google Drive)
# 👤 AUTOR: Marcos Barbosa - MoveRight (c)
# ⚙️ FUNÇÃO: Roteador Central, Segurança, Dashboard Principal e Temas.
# ==============================================================================

import streamlit as st
from utils.logger import configurar_logging
configurar_logging()

st.set_page_config(
    page_title="Esporte e Saúde - Gestão",
    layout="wide",
    page_icon="🏃‍♂️",
    initial_sidebar_state="expanded",
)

# ── Esconde "Running..." e toolbar antes de qualquer render (incl. tela de login) ──
st.markdown(
    "<style>"
    "[data-testid='stStatusWidget']{display:none !important;}"
    "[data-testid='stToolbar']{display:none !important;}"
    "[data-testid='stAppToolbar']{display:none !important;}"
    "[data-testid='stDecoration']{display:none !important;}"
    "[data-testid='stSpinnerContainer']{display:none !important;}"
    "[data-testid='stConnectionStatus']{display:none !important;}"
    ".stToolbar{display:none !important;}"
    ".stAppToolbar{display:none !important;}"
    "div[class*='StatusWidget']{display:none !important;}"
    "</style>",
    unsafe_allow_html=True,
)

import datetime
import time
import pandas as pd
import re
import random
import urllib.parse
import math

from database import (
    get_agendamentos_pendentes,
    autenticar_usuario,
    cadastrar_usuario_sistema,
    recuperar_senha_usuario,
    get_template_seguro_db,
    load_frequencia_ultima_presenca,
    load_atestados_vencimento,
    load_total_presencas_todos,
    ADMIN_MASTER,
    supabase,
)


# ==============================================================================
# 🎨 SELETOR DINÂMICO DE TEMA E FERRAMENTAS GLOBAIS
# ==============================================================================
def injetar_css_tema():
    """Injeta CSS do tema ativo com .stApp como prefixo para superar especificidade do CSS base."""
    if "tema_operador" not in st.session_state:
        st.session_state.tema_operador = "Claro"

    # Esconde o indicador "Running..." (status widget) do Streamlit em todas as telas
    st.markdown(
        "<style>[data-testid='stStatusWidget']{display:none !important;}</style>",
        unsafe_allow_html=True,
    )

    if st.session_state.tema_operador == "Escuro":
        st.markdown("""
<style>
/* ════════════════════════════════════════════════════════════════
   TEMA ESCURO v3 — prefixo .stApp garante especificidade máxima
════════════════════════════════════════════════════════════════ */

/* ── Fundos ─────────────────────────────────────────────────── */
.stApp                                        { background: #0E1117 !important; }
.stApp .stSidebar                             { background: #1E293B !important; }
.stApp div[data-testid="stVerticalBlockBorderWrapper"] {
    background: #1E293B !important;
    border-color: #334155 !important;
}

/* ── Texto global ─────────────────────────────────────────────── */
.stApp p, .stApp span, .stApp li,
.stApp h1, .stApp h2, .stApp h3,
.stApp h4, .stApp h5, .stApp h6,
.stApp label,
.stApp div[data-testid="stMarkdownContainer"] p { color: #F8FAFC !important; }
.stApp small, .stApp .stCaption p             { color: #94A3B8 !important; }

/* ── Nav radiogroup ───────────────────────────────────────────── */
.stApp div[role="radiogroup"]                 { background: #1E293B !important; border-color: #334155 !important; }
.stApp div[role="radiogroup"]::before         { color: #3B82F6 !important; }
.stApp div[role="radiogroup"] label p         { color: #94A3B8 !important; }
.stApp div[role="radiogroup"] label:hover     { background: #334155 !important; }
.stApp div[role="radiogroup"] label[data-checked="true"]   { background: #3B82F6 !important; }
.stApp div[role="radiogroup"] label[data-checked="true"] p { color: #FFFFFF !important; }

/* ── Inputs / Selects / Textareas ─────────────────────────────── */
.stApp div[data-baseweb="input"] > div,
.stApp div[data-baseweb="select"] > div,
.stApp div[data-baseweb="textarea"] > div {
    background: #1E293B !important;
    border-color: #334155 !important;
}
.stApp div[data-baseweb="input"] input,
.stApp div[data-baseweb="textarea"] textarea,
.stApp div[data-baseweb="select"] input,
.stApp div[data-baseweb="select"] [data-testid="stSelectboxValue"],
.stApp div[data-baseweb="select"] [data-testid="stSelectboxLabel"] { color: #F8FAFC !important; }

/* ── BOTÕES PRIMARY ───────────────────────────────────────────── */
.stApp button[kind="primary"],
.stApp button[kind="primaryFormSubmit"],
.stApp button[data-testid="stBaseButton-primary"],
.stApp button[data-testid="stBaseButton-primaryFormSubmit"] {
    background: linear-gradient(135deg,#0056b3 0%,#0072e5 100%) !important;
    color: #FFFFFF !important;
    border: none !important;
}
.stApp button[kind="primary"] p,
.stApp button[kind="primary"] span,
.stApp button[kind="primaryFormSubmit"] p,
.stApp button[kind="primaryFormSubmit"] span,
.stApp button[data-testid="stBaseButton-primary"] p,
.stApp button[data-testid="stBaseButton-primaryFormSubmit"] p { color: #FFFFFF !important; }

/* ── BOTÕES SECONDARY ─────────────────────────────────────────── */
.stApp button[kind="secondary"],
.stApp button[kind="secondaryFormSubmit"],
.stApp button[data-testid="stBaseButton-secondary"],
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] {
    background: #1E293B !important;
    border: 1.5px solid #475569 !important;
    color: #F8FAFC !important;
}
.stApp button[kind="secondary"]:hover,
.stApp button[kind="secondaryFormSubmit"]:hover,
.stApp button[data-testid="stBaseButton-secondary"]:hover,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"]:hover {
    background: #334155 !important;
    border-color: #64748B !important;
    color: #93C5FD !important;
}
.stApp button[kind="secondary"] p,
.stApp button[kind="secondary"] span,
.stApp button[kind="secondaryFormSubmit"] p,
.stApp button[kind="secondaryFormSubmit"] span,
.stApp button[data-testid="stBaseButton-secondary"] p,
.stApp button[data-testid="stBaseButton-secondary"] span,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] p,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] span { color: #F8FAFC !important; }
.stApp button[kind="secondary"]:hover p,
.stApp button[kind="secondary"]:hover span,
.stApp button[data-testid="stBaseButton-secondary"]:hover p,
.stApp button[data-testid="stBaseButton-secondary"]:hover span { color: #93C5FD !important; }

/* ── BOTÕES TERTIARY ──────────────────────────────────────────── */
.stApp button[kind="tertiary"],
.stApp button[data-testid="stBaseButton-tertiary"] {
    background: transparent !important;
    border: none !important;
    color: #93C5FD !important;
}
.stApp button[kind="tertiary"] p,
.stApp button[kind="tertiary"] span,
.stApp button[data-testid="stBaseButton-tertiary"] p,
.stApp button[data-testid="stBaseButton-tertiary"] span { color: #93C5FD !important; }

/* ── BOTÕES ICON ──────────────────────────────────────────────── */
.stApp button[kind="icon"],
.stApp button[data-testid="stBaseButton-icon"] {
    background: #1E293B !important;
    border-color: #334155 !important;
    color: #F8FAFC !important;
}
.stApp button[kind="icon"] p,
.stApp button[data-testid="stBaseButton-icon"] p { color: #F8FAFC !important; }

/* ── LINK BUTTONS ─────────────────────────────────────────────── */
.stApp a[data-testid="stLinkButton"],
.stApp .stLinkButton a {
    background: #1E293B !important;
    border: 1.5px solid #475569 !important;
    color: #F8FAFC !important;
}
.stApp a[data-testid="stLinkButton"]:hover,
.stApp .stLinkButton a:hover { background: #334155 !important; color: #93C5FD !important; }
.stApp a[data-testid="stLinkButton"] p,
.stApp a[data-testid="stLinkButton"] span,
.stApp .stLinkButton a p,
.stApp .stLinkButton a span { color: #F8FAFC !important; }

/* ── SORT HEADER (cabeçalhos clicáveis do grid) ───────────────── */
.stApp .sort-header button {
    background: transparent !important;
    border: none !important;
    color: #CBD5E1 !important;
    font-weight: 700 !important;
}
.stApp .sort-header button:hover { color: #60A5FA !important; }
.stApp .sort-header button p,
.stApp .sort-header button span { color: inherit !important; }

/* ── TABS ─────────────────────────────────────────────────────── */
.stApp button[data-baseweb="tab"]                           { color: #94A3B8 !important; }
.stApp button[data-baseweb="tab"][aria-selected="true"]     { color: #60A5FA !important; border-color: #60A5FA !important; }

/* ── EXPANDERS ────────────────────────────────────────────────── */
.stApp details summary                        { color: #F8FAFC !important; }
.stApp details[data-testid="stExpander"]      { border-color: #334155 !important; }

/* ── MÉTRICAS ─────────────────────────────────────────────────── */
.stApp div[data-testid="stMetricValue"]       { color: #F8FAFC !important; }
.stApp div[data-testid="stMetricLabel"] p     { color: #94A3B8 !important; }
.stApp div[data-testid="stMetricDelta"]       { color: #94A3B8 !important; }

/* ── ALERTAS / INFO / WARNING ─────────────────────────────────── */
.stApp div[data-testid="stAlert"] p           { color: inherit !important; }
</style>
""", unsafe_allow_html=True)

    else:
        # Tema Claro — especificidade .stApp garante contraste sobre qualquer override
        st.markdown("""
<style>
/* ════════════════════════════════════════════════════════════════
   TEMA CLARO v3 — contraste total em todos os tipos de botão
════════════════════════════════════════════════════════════════ */

/* ── BOTÕES SECONDARY ─────────────────────────────────────────── */
.stApp button[kind="secondary"],
.stApp button[kind="secondaryFormSubmit"],
.stApp button[data-testid="stBaseButton-secondary"],
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] {
    background: #F8FAFC !important;
    border: 1.5px solid #E2E8F0 !important;
    color: #1E293B !important;
}
.stApp button[kind="secondary"]:hover,
.stApp button[kind="secondaryFormSubmit"]:hover,
.stApp button[data-testid="stBaseButton-secondary"]:hover,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"]:hover {
    background: #EEF2FF !important;
    border-color: #C7D2FE !important;
    color: #0056b3 !important;
}
.stApp button[kind="secondary"] p,
.stApp button[kind="secondary"] span,
.stApp button[kind="secondaryFormSubmit"] p,
.stApp button[kind="secondaryFormSubmit"] span,
.stApp button[data-testid="stBaseButton-secondary"] p,
.stApp button[data-testid="stBaseButton-secondary"] span,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] p,
.stApp button[data-testid="stBaseButton-secondaryFormSubmit"] span { color: #1E293B !important; }
.stApp button[kind="secondary"]:hover p,
.stApp button[data-testid="stBaseButton-secondary"]:hover p { color: #0056b3 !important; }

/* ── BOTÕES TERTIARY ──────────────────────────────────────────── */
.stApp button[kind="tertiary"],
.stApp button[data-testid="stBaseButton-tertiary"] { color: #0056b3 !important; }
.stApp button[kind="tertiary"] p,
.stApp button[data-testid="stBaseButton-tertiary"] p { color: #0056b3 !important; }

/* ── LINK BUTTONS ─────────────────────────────────────────────── */
.stApp a[data-testid="stLinkButton"],
.stApp .stLinkButton a {
    background: #F8FAFC !important;
    border: 1.5px solid #E2E8F0 !important;
    color: #1E293B !important;
}
.stApp a[data-testid="stLinkButton"]:hover,
.stApp .stLinkButton a:hover {
    background: #EEF2FF !important;
    border-color: #C7D2FE !important;
    color: #0056b3 !important;
}
.stApp a[data-testid="stLinkButton"] p,
.stApp a[data-testid="stLinkButton"] span,
.stApp .stLinkButton a p,
.stApp .stLinkButton a span { color: #1E293B !important; }

/* ── SORT HEADER (cabeçalhos clicáveis do grid) ───────────────── */
.stApp .sort-header button {
    background: transparent !important;
    border: none !important;
    color: #0F172A !important;
    font-weight: 700 !important;
}
.stApp .sort-header button:hover { color: #1D4ED8 !important; }
.stApp .sort-header button p,
.stApp .sort-header button span { color: inherit !important; }
</style>
""", unsafe_allow_html=True)


def _toggle_tema_ui(key_suffix=""):
    """Widget compacto de toggle de tema — reutilizável no login e no app."""
    tema_escolhido = st.selectbox(
        "🌗",
        ["☀️ Claro", "🌙 Escuro"],
        index=0 if st.session_state.tema_operador == "Claro" else 1,
        label_visibility="collapsed",
        key=f"tema_sel_{key_suffix}",
    )
    novo = "Claro" if tema_escolhido.startswith("☀️") else "Escuro"
    if novo != st.session_state.tema_operador:
        st.session_state.tema_operador = novo
        st.rerun()


def renderizar_seletor_tema():
    """Renderiza Drive + Prestação de Contas + Seletor de Tema no topo do app."""
    col_vazia, col_drive, col_prest, col_tema = st.columns(
        [3.0, 2.5, 2.5, 2], vertical_alignment="center"
    )
    with col_drive:
        st.link_button(
            "📂 Abrir Google Drive",
            "https://drive.google.com/drive/u/7/my-drive",
            use_container_width=True,
            help="Acesse a pasta da nuvem para gerir as fotografias.",
        )
    with col_prest:
        st.link_button(
            "📁 Prestação de Contas",
            "https://drive.google.com/drive/folders/1TWP0Q3nwKpsDKmjWUrFC6uKBuvfmGpFC",
            use_container_width=True,
            help="Abra a pasta de Prestação de Contas para salvar os PDFs gerados.",
        )
    with col_tema:
        _toggle_tema_ui(key_suffix="app")


# ==============================================================================
# 🚪 ROTEADOR PÚBLICO COM "BOTÃO DE VOLTAR" E VALIDADOR DE QR CODE
# ==============================================================================
rota = st.query_params.get("rota")
if rota in ["inscricao", "pesquisa", "validar"]:
    st.markdown(
        "<style>#MainMenu, header, footer {visibility: hidden;} .block-container {padding-top: 1rem !important;}</style>",
        unsafe_allow_html=True,
    )

    if rota != "validar":
        col_back, _ = st.columns([1, 4])
        with col_back:
            if st.button("⬅️ Voltar ao Início", use_container_width=True):
                st.query_params.clear()
                st.rerun()
        st.markdown(
            "<hr style='margin-top: 0; margin-bottom: 20px; border-top: 2px solid #E2E8F0;'>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

    if rota == "inscricao":
        from views.inscricao_publica_view import tela_inscricao_publica_move_right

        tela_inscricao_publica_move_right()
    elif rota == "pesquisa":
        from views.pesquisa_satisfacao_view import tela_pesquisa_satisfacao_move_right

        tela_pesquisa_satisfacao_move_right()
    elif rota == "validar":
        from views.validador_view import tela_validador_publico

        tela_validador_publico()
    st.stop()


# ==============================================================================
# 📦 FUNÇÕES CACHED (nível de módulo — nunca dentro de condicionais)
# ==============================================================================
@st.cache_data(ttl=3600, show_spinner=False)
def load_niver_geral():
    try:
        from database import buscar_alunos_geral
        df = buscar_alunos_geral("")
        df["dt"] = pd.to_datetime(df["data_nascimento"], errors="coerce")
        df = df.dropna(subset=["dt"]).copy()
        df["dia"] = df["dt"].dt.day
        df["mes"] = df["dt"].dt.month
        return df
    except Exception:
        return pd.DataFrame()


# ==============================================================================
# 🛡️ MÓDULO DE SEGURANÇA E SESSÃO (8 HORAS)
# ==============================================================================
def gerar_captcha():
    st.session_state.captcha_n1 = random.randint(1, 10)
    st.session_state.captcha_n2 = random.randint(1, 10)
    st.session_state.captcha_result = (
        st.session_state.captcha_n1 + st.session_state.captcha_n2
    )


def inicializar_sessao():
    chaves = {
        "usuario_logado": False,
        "usuario_nome": "",
        "usuario_email": "",
        "perfil": "Visitante",
        "menu_atual": "Principal",
        "auth_tab": "in",
        "aluno_prontuario": None,
        "ultimo_acesso": time.time(),
        "admin_liberado": False,
    }
    for k, v in chaves.items():
        if k not in st.session_state:
            st.session_state[k] = v
    if "captcha_result" not in st.session_state:
        gerar_captcha()


inicializar_sessao()


def _ebi_base_url():
    """URL pública do app (https://host) para montar links acionáveis no e-mail."""
    try:
        h = st.context.headers.get("host", "")
        if h:
            return f"https://{h}"
    except Exception:
        pass
    return ""


# ── Deep-link interno: links acionáveis vindos do Email BI ────────────────────
# (?ir=freq&d=YYYY-MM-DD  → tela de Frequência | ?ir=ficha&id=<id> → ficha do aluno)
_ir_dl = st.query_params.get("ir")
if _ir_dl in ("freq", "ficha", "triagem") and "_pending_deeplink" not in st.session_state:
    st.session_state["_pending_deeplink"] = {
        "ir": _ir_dl,
        "d": st.query_params.get("d", ""),
        "id": st.query_params.get("id", ""),
    }
    try:
        st.query_params.clear()
    except Exception:
        pass

if st.session_state.usuario_logado:
    if time.time() - st.session_state.ultimo_acesso > 28800:
        st.session_state.clear()
        st.session_state.alerta_expiracao = "⚠️ A sua sessão expirou por medida de segurança após um longo período de inatividade. Por favor, acesse novamente."
        st.rerun()

    st.session_state.ultimo_acesso = time.time()

    # ── Aplica deep-link interno após autenticação ─────────────────────────
    _dl = st.session_state.pop("_pending_deeplink", None)
    if _dl:
        if _dl.get("ir") == "freq":
            st.session_state.menu_atual = "Frequência"
            _d_dl = _dl.get("d", "")
            if _d_dl:
                try:
                    st.session_state["_freq_data_alvo"] = datetime.date.fromisoformat(_d_dl[:10])
                except Exception:
                    pass
            st.rerun()
        elif _dl.get("ir") == "ficha" and _dl.get("id"):
            try:
                from database import buscar_aluno_por_id
                _al_dl = buscar_aluno_por_id(_dl["id"])
            except Exception:
                _al_dl = None
            if _al_dl:
                st.session_state.aluno_prontuario = _al_dl
                st.session_state.origem_prontuario = "Principal"
                st.session_state.menu_atual = "Portal do Aluno"
            else:
                st.session_state["_deeplink_erro"] = "⚠️ Aluno do link não encontrado."
            st.rerun()
        elif _dl.get("ir") == "triagem":
            st.session_state.aluno_prontuario = None
            st.session_state.menu_atual = "Portal do Aluno"
            st.session_state["_abrir_triagem"] = True
            st.rerun()

    if st.session_state.get("_deeplink_erro"):
        st.warning(st.session_state.pop("_deeplink_erro"))

    # ── Email BI: verificação de envio agendado (1x por sessão) ────────────
    if not st.session_state.get("_ebi_checado"):
        st.session_state["_ebi_checado"] = True
        try:
            from utils.email_relatorio_config import (
                get_schedules, schedule_to_cfg, marcar_envio_realizado_schedule,
                get_config_ebi, verificar_e_marcar_envio_realizado,
            )
            from utils.email_relatorio import enviar_relatorio_bi
            from utils.identidade import get_config as _gid_ebi
            _nome_org_ebi = _gid_ebi().get("nome_organizacao", "Instituto Muda Brasil")
            _hoje_ebi = datetime.date.today()
            _schedules = get_schedules()
            if _schedules:
                # Multi-schedule: itera por todos os pacotes habilitados
                for _sched in _schedules:
                    if not _sched.get("habilitado"):
                        continue
                    if not _sched.get("emails_destino"):
                        continue
                    _prox_s = _sched.get("proximo_envio", "")
                    _devido_s = not _prox_s
                    if not _devido_s:
                        try:
                            _devido_s = datetime.date.fromisoformat(str(_prox_s)[:10]) <= _hoje_ebi
                        except Exception:
                            _devido_s = True
                    if _devido_s:
                        _cfg_s = schedule_to_cfg(_sched)
                        _ok_s, _ = enviar_relatorio_bi(_cfg_s, _nome_org_ebi, _ebi_base_url())
                        if _ok_s:
                            marcar_envio_realizado_schedule(_sched["id"], _cfg_s)
            else:
                # Fallback legado (configuracoes_sistema) enquanto não há schedules
                _ebi_cfg = get_config_ebi()
                _prox = _ebi_cfg.get("proximo_envio", "")
                _devido = not _prox
                if not _devido:
                    try:
                        _devido = datetime.date.fromisoformat(str(_prox)[:10]) <= _hoje_ebi
                    except Exception:
                        _devido = True
                if _devido and _ebi_cfg.get("habilitado") and _ebi_cfg.get("emails_destino"):
                    _ok_ebi, _ = enviar_relatorio_bi(_ebi_cfg, _nome_org_ebi, _ebi_base_url())
                    if _ok_ebi:
                        verificar_e_marcar_envio_realizado(_ebi_cfg)
        except Exception:
            pass

    # ── Alertas de Ausência (WhatsApp/Z-API): verificação diária (1x por sessão)
    if not st.session_state.get("_alerta_ausencia_checado"):
        st.session_state["_alerta_ausencia_checado"] = True
        try:
            from database import get_config_valor as _gcv_aa
            if _gcv_aa("alerta_ausencia_habilitado", "0") == "1":
                from utils.niver_automatico import disparar_alertas_ausencia as _daa
                _daa()
        except Exception:
            pass

# ==============================================================================
# 🎨 CSS PRIME — MINIMALISTA & EXCELÊNCIA (TEMA CLARO BASE)
# ==============================================================================
st.markdown(
    """
<style>
/* ── BASE ──────────────────────────────────────────────────────────────────── */
#MainMenu, footer { visibility: hidden; }
[data-testid="stStatusWidget"]         { display: none !important; }
/* stHeader: transparente com altura natural — o collapsedControl precisa de área clicável */
[data-testid="stHeader"] {
    background: transparent !important;
    border-bottom: none !important;
    box-shadow: none !important;
}
/* Esconde apenas o conteúdo indesejado dentro do header */
[data-testid="stToolbar"]              { display: none !important; }
[data-testid="stAppToolbar"]           { display: none !important; }
[data-testid="stDecoration"]           { display: none !important; }
[data-testid="stSpinnerContainer"]     { display: none !important; }
[data-testid="stConnectionStatus"]     { display: none !important; }
[data-testid="stAppRunningIndicator"]  { display: none !important; }
[data-testid="stMainMenuPopover"]      { display: none !important; }
[data-testid="stMainMenu"]             { display: none !important; }
[data-testid="stStatusWidget"]         { display: none !important; }
.stAppToolbar                          { display: none !important; }
.stToolbar                             { display: none !important; }
#stDecoration                          { display: none !important; }
div[class*="StatusWidget"]             { display: none !important; }
div[class*="AppRunning"]               { display: none !important; }
div[class*="appRunning"]               { display: none !important; }
div[class*="runningIndicator"]         { display: none !important; }
div[class*="RunningIndicator"]         { display: none !important; }
/* collapsedControl — botão de reabrir sidebar — sempre visível e clicável */
[data-testid="collapsedControl"] {
    display: flex !important;
    pointer-events: all !important;
    z-index: 999 !important;
}
.block-container {
    padding-top: 0 !important;
    padding-left: 1.5rem !important;
    padding-right: 1.5rem !important;
    max-width: 100% !important;
    padding-bottom: 74px !important;
}
.stApp {
    background: linear-gradient(160deg,#EEF2FF 0%,#E8EEF8 45%,#F0F4FF 100%) !important;
}

/* ── CARD DO LOGIN ──────────────────────────────────────────────────────────── */
div[data-testid="column"] > div[data-testid="stVerticalBlockBorderWrapper"] {
    background: #FFFFFF !important;
    border-radius: 20px !important;
    box-shadow: 0 28px 56px rgba(10,37,64,.13), 0 6px 16px rgba(10,37,64,.07) !important;
    border: 1px solid rgba(226,232,240,.8) !important;
    border-top: none !important;
    padding: 0 !important;
    overflow: hidden !important;
}
div[data-testid="column"] > div[data-testid="stVerticalBlockBorderWrapper"]
    div[data-testid="stVerticalBlockBorderWrapper"] {
    background: transparent !important; border: none !important;
    box-shadow: none !important; padding: 0 !important; margin: 0 !important;
}

/* ── INPUTS ─────────────────────────────────────────────────────────────────── */
div[data-baseweb="input"] > div, div[data-baseweb="select"] > div {
    border-radius: 10px !important;
    border: 1.5px solid #E2E8F0 !important;
    background: #F8FAFC !important;
    transition: all .2s ease !important;
}
div[data-baseweb="input"] > div:focus-within, div[data-baseweb="select"] > div:focus-within {
    border-color: #0056b3 !important;
    background: #FFFFFF !important;
    box-shadow: 0 0 0 3px rgba(0,86,179,.08) !important;
}
div[data-baseweb="input"] input { font-size: 14px !important; padding: 10px 14px !important; }

/* ── BOTÕES PRIMARY ─────────────────────────────────────────────────────────── */
button[kind="primaryFormSubmit"], button[kind="primary"] {
    background: linear-gradient(135deg,#0056b3 0%,#0072e5 100%) !important;
    color: white !important; border: none !important;
    padding: 18px 15px !important; font-weight: 700 !important;
    font-size: 14px !important; border-radius: 10px !important;
    text-transform: uppercase; letter-spacing: 1.2px;
    width: 100% !important; transition: all .25s ease !important;
    box-shadow: 0 4px 14px rgba(0,86,179,.28) !important;
    margin-top: 6px !important;
}
button[kind="primaryFormSubmit"]:hover, button[kind="primary"]:hover {
    background: linear-gradient(135deg,#004494 0%,#0056b3 100%) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,86,179,.34) !important;
}

/* ── BOTÕES SECONDARY ───────────────────────────────────────────────────────── */
button[kind="secondary"] {
    background: #F8FAFC !important; border: 1.5px solid #E2E8F0 !important;
    color: #475569 !important; border-radius: 10px !important;
    font-weight: 600 !important; font-size: 13px !important;
    transition: all .2s ease !important;
}
button[kind="secondary"]:hover {
    background: #EEF2FF !important; border-color: #C7D2FE !important;
    color: #0056b3 !important;
}

/* ── CAPTCHA ────────────────────────────────────────────────────────────────── */
.captcha-box {
    display: flex; align-items: center; justify-content: center;
    background: #F0F7FF; border: 1.5px solid #BFDBFE;
    border-radius: 10px; height: 44px;
    color: #1E3A8A; font-weight: 800; font-size: 14px; letter-spacing: .5px;
}

/* ── LINKS PÚBLICOS ─────────────────────────────────────────────────────────── */
.pub-pill {
    display: flex; align-items: center; justify-content: center;
    gap: 5px; padding: 9px 12px;
    background: #F8FAFC; border: 1px solid #E2E8F0;
    color: #475569; border-radius: 8px; font-size: 12px; font-weight: 600;
    text-decoration: none; transition: all .2s ease;
}
.pub-pill:hover { background: #EEF2FF; color: #0056b3; border-color: #C7D2FE; }

/* ── RODAPÉ FIXO ────────────────────────────────────────────────────────────── */
.rodape-prime {
    position: fixed; bottom: 0; left: 0; width: 100%;
    background: rgba(10,37,64,.97); backdrop-filter: blur(8px);
    color: rgba(255,255,255,.65); text-align: center;
    padding: 9px 20px; z-index: 999; font-size: 11.5px;
    border-top: 1px solid rgba(255,255,255,.07);
}
.rodape-prime strong { color: #fff; }
.rodape-prime a { color: #94A3B8; text-decoration: none; margin: 0 6px; transition: .2s; }
.rodape-prime a:hover { color: #60A5FA; }

/* ── BARRA NAV (RADIO PILLS) ────────────────────────────────────────────────── */
div[role="radiogroup"] {
    background: #FFFFFF !important; padding: 7px 16px !important;
    border-radius: 14px !important; gap: 3px !important; align-items: center !important;
    box-shadow: 0 2px 8px rgba(0,0,0,.05) !important; border: 1px solid #E2E8F0 !important;
    margin-bottom: 16px !important; margin-top: 0 !important;
    justify-content: flex-end !important;
    flex-wrap: nowrap !important;
    overflow-x: auto !important;
    min-height: 44px !important;
}
div[role="radiogroup"]::before {
    content: "🏃 IMBRA"; color: #0056b3; font-size: 16px; font-weight: 900;
    margin-right: auto; letter-spacing: -.5px;
}
div[role="radiogroup"] label {
    background: transparent !important; border: none !important;
    border-radius: 8px !important; padding: 5px 10px !important;
    transition: all .18s ease !important; margin: 0 !important;
}
div[role="radiogroup"] label p {
    color: #64748B !important; font-weight: 600 !important;
    font-size: 12px !important; margin: 0 !important;
}
div[role="radiogroup"] label:hover { background: #F1F5F9 !important; }
div[role="radiogroup"] label[data-checked="true"] {
    background: #0056b3 !important;
    box-shadow: 0 2px 8px rgba(0,86,179,.22) !important;
}
div[role="radiogroup"] label[data-checked="true"] p {
    color: #FFFFFF !important; font-weight: 700 !important;
}

/* ── ATALHOS DO DASHBOARD ───────────────────────────────────────────────────── */
.stButton button, .stLinkButton a {
    border-radius: 12px !important; font-weight: 700 !important;
    font-size: 13px !important; letter-spacing: .2px;
}
</style>
""",
    unsafe_allow_html=True,
)

# CSS do tema injeta SEMPRE (antes do check de login)
injetar_css_tema()

# ==============================================================================
# 🔐 PORTAL DE ACESSO — PRIME
# ==============================================================================
if not st.session_state.usuario_logado:
    # Toggle de tema no canto superior direito do login
    _cl, _ct = st.columns([6, 1])
    with _ct:
        _toggle_tema_ui(key_suffix="login")

    st.markdown("<div style='min-height:3vh;'></div>", unsafe_allow_html=True)
    _, col_c, _ = st.columns([1, 1.05, 1])

    with col_c:
        if "alerta_expiracao" in st.session_state:
            st.warning(st.session_state.pop("alerta_expiracao"))

        with st.container(border=True):
            # ── HEADER ESCURO COM LOGO ──────────────────────────────────────
            from utils.identidade import (
                get_config as _gc_l,
                get_logo_data_url as _gld_l,
            )

            _cfg_l = _gc_l()
            _logo_b64 = _gld_l(_cfg_l.get("logo_principal", "logo-imbra.png"))
            _logo_html = (
                f'<img src="{_logo_b64}" style="height:52px;object-fit:contain;'
                f'filter:brightness(0) invert(1);margin-bottom:8px;">'
                if _logo_b64
                else '<div style="font-size:38px;margin-bottom:8px;">🏃‍♂️</div>'
            )
            st.markdown(
                f"""<div style="background:linear-gradient(135deg,#0A2540 0%,#1a3a5c 100%);
                              padding:26px 32px 22px;text-align:center;margin:-1px -1px 0;">
                    {_logo_html}
                    <h2 style="color:#FFFFFF;margin:0;font-size:13px;font-weight:800;
                               letter-spacing:.2px;line-height:1.3;white-space:nowrap;">
                        {_cfg_l.get("titulo_projeto", "ESPORTE E SAÚDE NA COMUNIDADE")}
                    </h2>
                    <p style="color:rgba(255,255,255,.72);font-size:12px;margin:6px 0 2px;
                              font-weight:500;letter-spacing:.1px;">
                        Entre com suas credenciais institucionais
                    </p>
                    <p style="color:rgba(255,255,255,.38);font-size:10px;margin:2px 0 0;
                              text-transform:uppercase;letter-spacing:2.5px;">
                        Gestão Inteligente MoveRight®
                    </p>
                </div>""",
                unsafe_allow_html=True,
            )

            # ── CORPO DO CARD ───────────────────────────────────────────────
            st.markdown("<div style='padding:22px 28px 20px;'>", unsafe_allow_html=True)

            if st.session_state.auth_tab == "in":
                with st.form("login_form"):
                    email = st.text_input(
                        "E-MAIL",
                        placeholder="exemplo@mudabrasil.org",
                        key="l_email",
                    )
                    senha = st.text_input(
                        "SENHA",
                        type="password",
                        placeholder="••••••••",
                        key="l_pwd",
                    )
                    col_cap, col_resp = st.columns([1.25, 1])
                    with col_cap:
                        st.markdown(
                            f"<div class='captcha-box'>🛡️ "
                            f"{st.session_state.captcha_n1} + "
                            f"{st.session_state.captcha_n2} = ?</div>",
                            unsafe_allow_html=True,
                        )
                    with col_resp:
                        resp = st.text_input(
                            "Resultado",
                            label_visibility="collapsed",
                            placeholder="Resposta",
                            key="l_cap",
                        )
                    btn_login = st.form_submit_button(
                        "ENTRAR  →", type="primary", use_container_width=True
                    )

                if btn_login:
                    if resp.strip() == str(st.session_state.captcha_result):
                        ok, user = autenticar_usuario(email, senha)
                        if ok:
                            st.session_state.usuario_logado = True
                            st.session_state.usuario_nome = user.get("nome")
                            st.session_state.usuario_email = email
                            st.session_state.email_usuario = email
                            st.session_state.usuario_id = user.get("id", "")
                            st.session_state.perfil = (
                                "SuperAdmin"
                                if email.lower() == ADMIN_MASTER.lower()
                                else "Admin"
                            )
                            # Limpa cache de permissões ao fazer login
                            st.session_state.pop("_menu_perms_cache", None)
                            st.session_state.pop("_menu_perms_version", None)
                            st.rerun()
                        else:
                            st.error("❌ Credenciais inválidas.")
                            gerar_captcha()
                    else:
                        st.error("❌ Verificação incorreta.")
                        gerar_captcha()
                        time.sleep(1)
                        st.rerun()

                st.markdown("<div style='height:4px;'></div>", unsafe_allow_html=True)
                cb1, cb2 = st.columns(2)
                with cb1:
                    if st.button(
                        "👤 Novo Operador", use_container_width=True, type="secondary"
                    ):
                        st.session_state.auth_tab = "up"
                        st.rerun()
                with cb2:
                    if st.button(
                        "🔑 Recuperar Senha", use_container_width=True, type="secondary"
                    ):
                        st.session_state.auth_tab = "forgot"
                        st.rerun()

                st.markdown(
                    "<p style='text-align:center;color:#94A3B8;font-size:11px;"
                    "margin:10px 0 2px;'>🔒 Acesso seguro — dados protegidos</p>"
                    "<hr style='border:none;border-top:1px solid #F1F5F9;margin:10px 0;'>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<a href="/?rota=pesquisa" target="_self" class="pub-pill">'
                    "⭐ Avaliar Projeto</a>",
                    unsafe_allow_html=True,
                )

            elif st.session_state.auth_tab == "up":
                if st.button("← Voltar", type="secondary"):
                    st.session_state.auth_tab = "in"
                    st.rerun()
                st.info("Apenas coordenadores podem criar contas de acesso.", icon="ℹ️")
                with st.form("reg_form"):
                    n = st.text_input("Nome Completo")
                    e = st.text_input("E-mail Institucional")
                    p = st.text_input("Criar Senha", type="password")
                    if st.form_submit_button(
                        "CRIAR CONTA", type="primary", use_container_width=True
                    ):
                        sucesso, msg = cadastrar_usuario_sistema(n, e, p)
                        if sucesso:
                            st.success(msg)
                            time.sleep(1)
                            st.session_state.auth_tab = "in"
                            st.rerun()
                        else:
                            st.error(msg)

            elif st.session_state.auth_tab == "forgot":
                if st.button("← Voltar", type="secondary"):
                    st.session_state.auth_tab = "in"
                    st.rerun()
                st.markdown(
                    "<p style='color:#0056b3;font-weight:700;margin:6px 0 2px;'>"
                    "🔑 Recuperar Acesso</p>"
                    "<p style='color:#64748B;font-size:12px;margin:0 0 10px;'>"
                    "Enviaremos instruções para o e-mail registado.</p>",
                    unsafe_allow_html=True,
                )
                with st.form("forgot_form"):
                    email_rec = st.text_input(
                        "E-mail de acesso",
                        placeholder="exemplo@mudabrasil.org",
                        key="r_email",
                    )
                    if st.form_submit_button(
                        "ENVIAR INSTRUÇÕES", type="primary", use_container_width=True
                    ):
                        sucesso, msg = recuperar_senha_usuario(email_rec.strip())
                        if sucesso:
                            st.success(f"✅ {msg}")
                            time.sleep(3)
                            st.session_state.auth_tab = "in"
                            st.rerun()
                        else:
                            st.error(f"❌ {msg}")

            st.markdown("</div>", unsafe_allow_html=True)

    # ── Rodapé do Login ─────────────────────────────────────────────────────
    from utils.identidade import get_config as _gcfg_login

    _lcfg = _gcfg_login()
    _links = []
    if _lcfg.get("site"):
        _links.append(
            f'<a href="https://{_lcfg["site"]}" target="_blank">🌐 {_lcfg["site"]}</a>'
        )
    if _lcfg.get("instagram"):
        _links.append(
            f'<a href="https://instagram.com/{_lcfg["instagram"].lstrip("@")}"'
            f' target="_blank">📸 {_lcfg["instagram"]}</a>'
        )
    if _lcfg.get("cnpj"):
        _links.append(f"CNPJ: {_lcfg['cnpj']}")
    st.markdown(
        f'<div class="rodape-prime">'
        f"<strong>{_lcfg.get('nome_organizacao', 'Instituto Muda Brasil')}</strong>"
        f"&nbsp;·&nbsp;{' &nbsp;|&nbsp; '.join(_links)}"
        f"</div>",
        unsafe_allow_html=True,
    )
    st.stop()

# ==============================================================================
# 📅 CALENDÁRIO INSTITUCIONAL — Tela de gestão de Dias Sem Aula
# ==============================================================================

_SQL_ALERTA_AUSENCIA = """
CREATE TABLE IF NOT EXISTS alertas_ausencia_log (
  id           uuid DEFAULT gen_random_uuid() PRIMARY KEY,
  aluno_id     text NOT NULL,
  limiar_dias  integer NOT NULL,
  data_alerta  date NOT NULL DEFAULT CURRENT_DATE,
  sucesso      boolean NOT NULL DEFAULT true,
  criado_em    timestamptz NOT NULL DEFAULT now()
);
ALTER TABLE alertas_ausencia_log DISABLE ROW LEVEL SECURITY;
""".strip()

_SQL_MIGRACAO_EBI = """
CREATE TABLE IF NOT EXISTS email_bi_schedules (
  id               uuid         DEFAULT gen_random_uuid() PRIMARY KEY,
  nome             text         NOT NULL DEFAULT 'Pacote Principal',
  habilitado       boolean      NOT NULL DEFAULT false,
  frequencia       text         NOT NULL DEFAULT 'semanal',
  dia_semana       integer      NOT NULL DEFAULT 4,
  dia_mes          integer      NOT NULL DEFAULT 1,
  emails_destino   jsonb        NOT NULL DEFAULT '[]'::jsonb,
  modulos          jsonb        NOT NULL DEFAULT '{}'::jsonb,
  assunto_extra    text                  DEFAULT '',
  email_remetente  text                  DEFAULT '',
  email_senha_app  text                  DEFAULT '',
  base_url         text                  DEFAULT '',
  proximo_envio    date,
  ultimo_envio     timestamptz,
  total_envios     integer      NOT NULL DEFAULT 0,
  historico_envios jsonb        NOT NULL DEFAULT '[]'::jsonb,
  criado_em        timestamptz  NOT NULL DEFAULT now(),
  atualizado_em    timestamptz  NOT NULL DEFAULT now()
);

-- Desativa RLS (sistema interno — acesso via chave de serviço)
ALTER TABLE email_bi_schedules DISABLE ROW LEVEL SECURITY;
""".strip()


def _tela_email_bi():
    import pandas as _pd_ebi
    from utils.email_relatorio_config import (
        calcular_proximo_envio,
        get_schedules, salvar_schedule, excluir_schedule,
        migrar_legado_para_schedule, schedule_to_cfg,
        marcar_envio_realizado_schedule, _ler_schedules,
    )
    from utils.email_relatorio import enviar_relatorio_bi
    from database import get_emails_sistema

    # ── Cabeçalho ─────────────────────────────────────────────────────────
    st.markdown("""
        <div style='background:#EFF6FF;border-left:4px solid #1D4ED8;
                    padding:12px 16px;border-radius:6px;margin-bottom:18px;'>
            <strong style='color:#1E3A8A;'>📧 Email BI — Relatório Gerencial Automático</strong><br>
            <span style='color:#1D4ED8;font-size:13px;'>
                Configure <strong>pacotes independentes</strong> de envio por e-mail.
                Cada pacote tem seus próprios destinatários, módulos e agendamento —
                com histórico completo de envios.
            </span>
        </div>
    """, unsafe_allow_html=True)

    # ── Verificar tabela ───────────────────────────────────────────────────
    schedules = get_schedules()
    if schedules is None:
        st.error("⚠️ Tabela `email_bi_schedules` não encontrada no banco de dados.")
        st.markdown(
            "Execute o SQL abaixo no **Supabase Dashboard → SQL Editor** "
            "e depois clique em **Confirmar**:"
        )
        st.code(_SQL_MIGRACAO_EBI, language="sql")
        if st.button("✅ Já executei — recarregar", key="ebi_migr_ok", type="primary"):
            _ler_schedules.clear()
            st.rerun()
        return

    # ── Auto-migração legada (1x por sessão) ──────────────────────────────
    if not schedules and not st.session_state.get("_ebi_migr_done"):
        if migrar_legado_para_schedule():
            _ler_schedules.clear()
            schedules = get_schedules() or []
        st.session_state["_ebi_migr_done"] = True

    emails_sistema = get_emails_sistema()

    # ── STATE MACHINE: lista ↔ editor ─────────────────────────────────────
    edit_id = st.session_state.get("ebi_edit_id")  # None | "new" | uuid

    # ══════════════════════════════════════════════════════════════════════
    # EDITOR
    # ══════════════════════════════════════════════════════════════════════
    if edit_id is not None:
        dados = {}
        if edit_id != "new":
            dados = next((s for s in schedules if s["id"] == edit_id), {})

        mod_def = dados.get("modulos") or {}
        if isinstance(mod_def, str):
            import json as _jj
            try:
                mod_def = _jj.loads(mod_def)
            except Exception:
                mod_def = {}

        emails_ja = dados.get("emails_destino") or []
        emails_sis_set = {u["email"] for u in emails_sistema}
        emails_externos_ja = [e for e in emails_ja if e not in emails_sis_set]

        titulo_ed = (
            f"📝 Editando: **{dados.get('nome', '—')}**"
            if edit_id != "new" else "📝 Novo Pacote de Envio"
        )
        st.markdown(f"### {titulo_ed}")

        with st.form(key=f"ebi_form_{edit_id}"):
            cn, ctog = st.columns([3, 2])
            nome_pct = cn.text_input(
                "Nome do pacote",
                value=dados.get("nome", "Pacote de Envio"),
                key=f"ebi_fn_{edit_id}",
            )
            habilitado = ctog.toggle(
                "Ativar envio automático",
                value=bool(dados.get("habilitado", False)),
                key=f"ebi_fh_{edit_id}",
                help="Quando ativo, o envio dispara automaticamente ao abrir o sistema na data agendada.",
            )

            # ── Agendamento ──────────────────────────────────────────────
            st.markdown("#### ⚙️ Agendamento")
            cfreq1, cfreq2 = st.columns(2)
            freq_opts = {"semanal": "Semanal", "quinzenal": "Quinzenal", "mensal": "Mensal"}
            freq_keys = list(freq_opts.keys())
            freq_atual = dados.get("frequencia", "semanal")
            frequencia = cfreq1.selectbox(
                "Frequência",
                options=freq_keys,
                index=freq_keys.index(freq_atual) if freq_atual in freq_keys else 0,
                format_func=lambda k: freq_opts[k],
                key=f"ebi_ff_{edit_id}",
            )
            dias_sem_nms = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
            dsv = int(dados.get("dia_semana") or 4)
            dmv = int(dados.get("dia_mes") or 1)
            if frequencia in ("semanal", "quinzenal"):
                dia_semana_sel = cfreq2.selectbox(
                    "Dia da semana",
                    options=list(range(7)),
                    index=dsv if 0 <= dsv <= 6 else 4,
                    format_func=lambda i: dias_sem_nms[i],
                    key=f"ebi_fds_{edit_id}",
                )
                dia_mes_sel = dmv
            else:
                dia_semana_sel = dsv
                dia_mes_sel = cfreq2.number_input(
                    "Dia do mês (1–28)", min_value=1, max_value=28,
                    value=dmv, step=1, key=f"ebi_fdm_{edit_id}",
                )

            # ── Destinatários ────────────────────────────────────────────
            st.markdown("#### 📬 Destinatários e Remetente")
            if emails_sistema:
                st.markdown("**E-mails cadastrados no sistema** *(marque os que devem receber este pacote):*")
                cols_em = st.columns(2)
                emails_sis_sel = []
                for idx_u, u in enumerate(emails_sistema):
                    col_u = cols_em[idx_u % 2]
                    lbl_u = f"{u['nome']} — {u['email']}"
                    chk_key = f"ebi_u_{edit_id}_{u['email'].replace('@','_at_').replace('.','_')}"
                    if col_u.checkbox(lbl_u, value=(u["email"] in emails_ja), key=chk_key):
                        emails_sis_sel.append(u["email"])
            else:
                emails_sis_sel = []
                st.caption("Nenhum usuário cadastrado no sistema ainda.")

            st.markdown("**E-mails adicionais** *(externos, um por linha):*")
            externos_txt = st.text_area(
                "",
                value="\n".join(emails_externos_ja),
                placeholder="emailexterno@exemplo.com\noutro@exemplo.com",
                height=75,
                label_visibility="collapsed",
                key=f"ebi_fext_{edit_id}",
            )

            # ── Remetente ────────────────────────────────────────────────
            cre1, cre2 = st.columns(2)
            remetente = cre1.text_input(
                "Gmail remetente",
                value=dados.get("email_remetente", ""),
                placeholder="seuemail@gmail.com",
                key=f"ebi_frem_{edit_id}",
            )
            senha_app = cre2.text_input(
                "Senha de app do Gmail",
                value=dados.get("email_senha_app", ""),
                type="password",
                key=f"ebi_fsen_{edit_id}",
                help="Gere em myaccount.google.com → Segurança → Senhas de app.",
            )
            base_url_cfg = st.text_input(
                "🔗 URL pública do sistema",
                value=dados.get("base_url", ""),
                placeholder="https://seusistema.onrender.com",
                key=f"ebi_furl_{edit_id}",
                help="Endereço público — usado nos botões de ação dentro do e-mail.",
            )

            # ── Módulos ──────────────────────────────────────────────────
            st.markdown("#### 🧩 Módulos do Relatório")
            cm1, cm2 = st.columns(2)
            mod_exec   = cm1.checkbox("📊 Painel Executivo (KPIs)",       value=bool(mod_def.get("executivo", True)),         key=f"ebi_me_{edit_id}")
            mod_ev     = cm1.checkbox("⚠️ Risco de Evasão",               value=bool(mod_def.get("evasao", True)),            key=f"ebi_mev_{edit_id}")
            mod_aud    = cm1.checkbox("📋 Auditoria de Cadastros",         value=bool(mod_def.get("auditoria", True)),         key=f"ebi_ma_{edit_id}")
            mod_nov    = cm1.checkbox("📥 Novos Cadastros (Aprovação)",    value=bool(mod_def.get("novos_cadastros", True)),   key=f"ebi_mn_{edit_id}")
            mod_freq   = cm2.checkbox("🏆 Frequência por Turma",           value=bool(mod_def.get("frequencia_turma", True)),  key=f"ebi_mft_{edit_id}")
            mod_dias   = cm2.checkbox("📅 Dias sem Registro",              value=bool(mod_def.get("dias_sem_registro", True)), key=f"ebi_mds_{edit_id}")
            mod_aniv   = cm2.checkbox("🎂 Aniversariantes da Semana",      value=bool(mod_def.get("aniversariantes", True)),   key=f"ebi_man_{edit_id}")
            mod_pres   = cm2.checkbox("📈 Presenças no Ano (por mês)",     value=bool(mod_def.get("presencas_mes", True)),     key=f"ebi_mpm_{edit_id}")
            mod_ates   = cm2.checkbox("🏥 Atestados a Vencer (30 dias)",  value=bool(mod_def.get("atestados_vencendo", False)), key=f"ebi_mat_{edit_id}")

            assunto_extra = st.text_input(
                "Texto extra no assunto (opcional)",
                value=dados.get("assunto_extra", ""),
                placeholder="Ex.: Unidade Centro",
                key=f"ebi_fass_{edit_id}",
            )

            st.markdown("---")
            cs1, cs2, cs3 = st.columns(3)
            btn_salvar   = cs1.form_submit_button("💾 Salvar Pacote",           type="primary", use_container_width=True)
            btn_teste    = cs2.form_submit_button("📨 Enviar Agora (Teste)",    use_container_width=True)
            btn_cancelar = cs3.form_submit_button("❌ Cancelar",                use_container_width=True)

        # ── Ações ────────────────────────────────────────────────────────
        if btn_cancelar:
            st.session_state.pop("ebi_edit_id", None)
            st.rerun()

        if btn_salvar or btn_teste:
            externos_lista = [
                e.strip()
                for linha in externos_txt.replace(",", "\n").splitlines()
                for e in [linha] if e.strip()
            ]
            todos_emails = list(dict.fromkeys(emails_sis_sel + externos_lista))
            cfg_calc = {"habilitado": habilitado, "frequencia": frequencia,
                        "dia_semana": int(dia_semana_sel), "dia_mes": int(dia_mes_sel)}
            proximo = calcular_proximo_envio(cfg_calc)
            payload = {
                "nome":            nome_pct.strip() or "Pacote de Envio",
                "habilitado":      habilitado,
                "frequencia":      frequencia,
                "dia_semana":      int(dia_semana_sel),
                "dia_mes":         int(dia_mes_sel),
                "emails_destino":  todos_emails,
                "modulos": {
                    "executivo":         mod_exec,
                    "evasao":            mod_ev,
                    "auditoria":         mod_aud,
                    "novos_cadastros":   mod_nov,
                    "frequencia_turma":  mod_freq,
                    "dias_sem_registro": mod_dias,
                    "aniversariantes":    mod_aniv,
                    "presencas_mes":      mod_pres,
                    "atestados_vencendo": mod_ates,
                },
                "assunto_extra":   assunto_extra.strip(),
                "email_remetente": remetente.strip(),
                "email_senha_app": senha_app.strip(),
                "base_url":        base_url_cfg.strip().rstrip("/"),
                "proximo_envio":   str(proximo),
            }
            if btn_salvar:
                sid = salvar_schedule(payload, edit_id if edit_id != "new" else None)
                if sid:
                    st.success(f"✅ Pacote **{payload['nome']}** salvo. Próximo envio: {proximo.strftime('%d/%m/%Y')}.")
                    st.session_state.pop("ebi_edit_id", None)
                    st.rerun()

            if btn_teste:
                if not todos_emails:
                    st.error("❌ Selecione ao menos um destinatário antes de testar.")
                else:
                    from utils.identidade import get_config as _gid_t
                    nome_org_t = _gid_t().get("nome_organizacao", "Instituto Muda Brasil")
                    cfg_t = schedule_to_cfg({**payload, "emails_destino": todos_emails})
                    with st.spinner("Gerando e enviando relatório de teste…"):
                        ok_t, msg_t = enviar_relatorio_bi(cfg_t, nome_org_t, _ebi_base_url())
                    if ok_t:
                        st.success(f"✅ {msg_t}")
                    else:
                        st.error(f"❌ {msg_t}")
        return

    # ══════════════════════════════════════════════════════════════════════
    # LISTAGEM DE PACOTES
    # ══════════════════════════════════════════════════════════════════════
    col_novo, _ = st.columns([2, 5])
    if col_novo.button("➕ Novo Pacote de Envio", type="primary", use_container_width=True, key="ebi_novo"):
        st.session_state["ebi_edit_id"] = "new"
        st.rerun()

    if not schedules:
        st.info("Nenhum pacote configurado ainda. Clique em **'➕ Novo Pacote'** para começar.")

    if schedules:
        st.markdown("---")
    _DIAS_NMS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    _FREQ_LBL = {"semanal": "Semanal", "quinzenal": "Quinzenal", "mensal": "Mensal"}

    for _s in schedules:
        _ativo   = bool(_s.get("habilitado", False))
        _icon    = "🟢" if _ativo else "⚪"
        _freq_l  = _FREQ_LBL.get(_s.get("frequencia", "semanal"), "—")
        _ds      = _s.get("dia_semana")
        _dm      = _s.get("dia_mes")
        if _s.get("frequencia", "semanal") in ("semanal", "quinzenal") and _ds is not None:
            _dia_l = _DIAS_NMS[int(_ds)] if 0 <= int(_ds) <= 6 else "—"
        else:
            _dia_l = f"dia {_dm}"
        _prox_l  = _s.get("proximo_envio") or "—"
        _total_l = _s.get("total_envios") or 0
        _ult_l   = (_s.get("ultimo_envio") or "")[:16].replace("T", " ") or "—"
        _n_em    = len(_s.get("emails_destino") or [])
        _sid     = _s["id"]

        with st.container(border=True):
            cL, cM1, cM2, cM3, cR = st.columns([3.5, 1.5, 1.5, 1.5, 2.5])
            cL.markdown(f"**{_s['nome']}** &nbsp; {_icon} {'Ativo' if _ativo else 'Inativo'}")
            cL.caption(f"✉️ {_n_em} destinatário(s) · {_freq_l} ({_dia_l})")
            cM1.metric("Próximo envio", _prox_l if _prox_l != "—" else "—")
            cM2.metric("Enviados", f"{_total_l}×")
            cM3.metric("Último", _ult_l[:10] if _ult_l != "—" else "—")

            ba, bt, bh, be = cR.columns(4)
            if ba.button("✏️", key=f"ebi_ed_{_sid}", help="Editar pacote"):
                st.session_state["ebi_edit_id"] = _sid
                st.rerun()
            if bt.button("📨", key=f"ebi_tst_{_sid}", help="Enviar agora (teste)"):
                st.session_state[f"ebi_test_{_sid}"] = True
                st.rerun()
            if bh.button("📋", key=f"ebi_hist_{_sid}", help="Ver histórico de envios"):
                _hk = f"ebi_show_hist_{_sid}"
                st.session_state[_hk] = not st.session_state.get(_hk, False)
                st.rerun()
            if be.button("🗑️", key=f"ebi_del_{_sid}", help="Excluir pacote"):
                st.session_state[f"ebi_conf_del_{_sid}"] = True
                st.rerun()

            # ── Confirmar exclusão ────────────────────────────────────────
            if st.session_state.get(f"ebi_conf_del_{_sid}"):
                st.warning(f"⚠️ Excluir **'{_s['nome']}'**? Esta ação não pode ser desfeita.")
                cd1, cd2 = st.columns(2)
                if cd1.button("✅ Confirmar exclusão", key=f"ebi_del_ok_{_sid}", type="primary"):
                    excluir_schedule(_sid)
                    st.session_state.pop(f"ebi_conf_del_{_sid}", None)
                    st.rerun()
                if cd2.button("Cancelar", key=f"ebi_del_cancel_{_sid}"):
                    st.session_state.pop(f"ebi_conf_del_{_sid}", None)
                    st.rerun()

            # ── Histórico ────────────────────────────────────────────────
            if st.session_state.get(f"ebi_show_hist_{_sid}"):
                _hist = _s.get("historico_envios") or []
                if isinstance(_hist, str):
                    import json as _jh
                    try:
                        _hist = _jh.loads(_hist)
                    except Exception:
                        _hist = []
                if not _hist:
                    st.caption("Nenhum envio registrado ainda.")
                else:
                    st.markdown(f"**📋 Histórico — {len(_hist)} envio(s) registrado(s):**")
                    _linhas_h = [
                        [str(i + 1), dt[:16].replace("T", " ")]
                        for i, dt in enumerate(reversed(_hist[-20:]))
                    ]
                    st.dataframe(
                        _pd_ebi.DataFrame(_linhas_h, columns=["#", "Data / Hora"]),
                        use_container_width=True, hide_index=True,
                    )

            # ── Envio de teste inline ─────────────────────────────────────
            if st.session_state.get(f"ebi_test_{_sid}"):
                st.session_state.pop(f"ebi_test_{_sid}")
                from utils.identidade import get_config as _gid_l
                _nome_l = _gid_l().get("nome_organizacao", "Instituto Muda Brasil")
                _cfg_l  = schedule_to_cfg(_s)
                with st.spinner(f"Enviando teste '{_s['nome']}'…"):
                    _ok_l, _msg_l = enviar_relatorio_bi(_cfg_l, _nome_l, _ebi_base_url())
                if _ok_l:
                    st.success(f"✅ {_msg_l}")
                else:
                    st.error(f"❌ {_msg_l}")

    # ══════════════════════════════════════════════════════════════════════
    # 📱 ALERTAS AUTOMÁTICOS DE AUSÊNCIA (WhatsApp / Z-API)
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("""
        <div style='background:#FFF7ED;border-left:4px solid #EA580C;
                    padding:12px 16px;border-radius:6px;margin-bottom:18px;'>
            <strong style='color:#9A3412;'>📱 Alertas Automáticos de Ausência (WhatsApp)</strong><br>
            <span style='color:#C2410C;font-size:13px;'>
                Dispara mensagem via Z-API quando um aluno ultrapassa <strong>30 dias</strong>
                ou <strong>60 dias</strong> sem presença registrada.
                Usa as mesmas credenciais Z-API configuradas em
                <strong>🔔 Auto Niver</strong>.
            </span>
        </div>
    """, unsafe_allow_html=True)

    from database import get_config_valor as _gcv_aa, set_config_valor as _scv_aa

    _aa_hab = _gcv_aa("alerta_ausencia_habilitado", "0") == "1"

    _col_aa1, _col_aa2 = st.columns([3, 2])
    _novo_aa_hab = _col_aa1.toggle(
        "Habilitar alertas automáticos de ausência",
        value=_aa_hab,
        key="aa_toggle_hab",
        help=(
            "Quando ativo, verifica diariamente (ao abrir o sistema) se algum aluno "
            "cruzou 30 ou 60 dias de ausência e envia WhatsApp automaticamente."
        ),
    )
    if _novo_aa_hab != _aa_hab:
        _scv_aa("alerta_ausencia_habilitado", "1" if _novo_aa_hab else "0")
        st.rerun()

    if _novo_aa_hab:
        st.success("✅ Ativo — o alerta é disparado automaticamente ao abrir o sistema (1× por sessão).")
    else:
        st.info("⚪ Inativo — nenhum alerta automático será enviado.")

    st.caption(
        "**Template 30 dias** → `evasao_60` · "
        "**Template 60 dias** → `evasao_80`   "
        "(edite os textos em ⚙️ Configurações → 💬 Mensagens)"
    )

    # ── Verificar se a tabela existe no banco ─────────────────────────────
    from database import get_ultimos_alertas_ausencia as _gua
    _hist_aa = _gua(limite=20)
    if _hist_aa is None:
        # tabela não existe ainda
        st.warning("⚠️ A tabela `alertas_ausencia_log` ainda não foi criada no banco de dados.")
        st.markdown("Execute o SQL abaixo no **Supabase Dashboard → SQL Editor** e recarregue:")
        st.code(_SQL_ALERTA_AUSENCIA, language="sql")
    else:
        # ── Disparar agora (teste manual) ─────────────────────────────────
        _col_btn1, _col_btn2, _ = st.columns([2, 2, 3])
        if _col_btn1.button("▶️ Disparar agora", key="aa_run_now", help="Executa a verificação imediatamente (modo manual)"):
            st.session_state["_aa_run_manual"] = True
            st.rerun()

        if st.session_state.pop("_aa_run_manual", False):
            with st.spinner("Verificando ausências e enviando alertas…"):
                try:
                    from utils.niver_automatico import disparar_alertas_ausencia as _daa_ui
                    _res_aa = _daa_ui()
                    if _res_aa:
                        _ok_aa  = [r for r in _res_aa if r.get("sucesso")]
                        _fail_aa = [r for r in _res_aa if not r.get("sucesso")]
                        if _ok_aa:
                            st.success(f"✅ {len(_ok_aa)} alerta(s) enviado(s) com sucesso.")
                        if _fail_aa:
                            st.warning(f"⚠️ {len(_fail_aa)} envio(s) falharam: " +
                                       ", ".join(r['nome'] for r in _fail_aa))
                    else:
                        st.info("Nenhum aluno elegível encontrado (Z-API inativo, ausências abaixo do limiar, ou alertas já enviados recentemente).")
                except Exception as _e_aa:
                    st.error(f"Erro ao disparar: {_e_aa}")

        # ── Histórico de alertas ───────────────────────────────────────────
        if _col_btn2.button("🔄 Atualizar histórico", key="aa_refresh_hist"):
            st.rerun()

        if _hist_aa:
            st.markdown(f"**📋 Últimos {len(_hist_aa)} alerta(s) registrado(s):**")
            import pandas as _pd_aa
            _df_hist_aa = _pd_aa.DataFrame([
                {
                    "#": i + 1,
                    "Data": str(r.get("data_alerta", ""))[:10],
                    "Limiar": f"{r.get('limiar_dias')} dias",
                    "Status": "✅ Enviado" if r.get("sucesso") else "❌ Falhou",
                    "Criado em": str(r.get("criado_em", ""))[:16].replace("T", " "),
                }
                for i, r in enumerate(_hist_aa)
            ])
            st.dataframe(_df_hist_aa, use_container_width=True, hide_index=True)
        else:
            st.caption("Nenhum alerta enviado ainda.")


def _tela_calendario_institucional():
    from database import (
        get_dias_sem_aula_periodo_df,
        registrar_dia_sem_aula,
        remover_dia_sem_aula,
        get_datas_comemorativas_bd,
    )

    st.markdown("""
        <div style='background:#F0FDF4;border-left:4px solid #16A34A;
                    padding:12px 16px;border-radius:6px;margin-bottom:18px;'>
            <strong style='color:#14532D;'>📅 Calendário Institucional</strong><br>
            <span style='color:#15803D;font-size:13px;'>
                Registre dias especiais: <strong>sem aula</strong> (reuniões, recessos,
                feriados locais) ou <strong>datas comemorativas com aula</strong>
                (aniversário da instituição, eventos festivos, etc.).<br>
                Dias sem aula são excluídos do alerta de <em>frequência pendente</em>
                nos relatórios. Datas comemorativas com aula ativam o
                badge festivo 🎉 e os balões 🎈 na tela de Frequência.
            </span>
        </div>
    """, unsafe_allow_html=True)

    hoje_cal = datetime.date.today()

    # chave rotativa para limpar o formulário após salvar
    fk = st.session_state.get("cal_form_key", 0)

    # ── Formulário de cadastro ─────────────────────────────────────────────
    st.markdown("### ➕ Registrar data especial")
    ca, cb = st.columns([2, 5])
    novo_dia_cal   = ca.date_input("Data:", value=hoje_cal, format="DD/MM/YYYY",
                                   key=f"cal_data_novo_{fk}")
    motivo_cal_txt = cb.text_input(
        "Motivo / Descrição:",
        placeholder="Ex: Reunião pedagógica, Aniversário IMBRA, Feriado municipal…",
        key=f"cal_motivo_{fk}",
    )

    # Checkboxes de tipo — em linha para economizar espaço vertical
    _cc1, _cc2, _cc3 = st.columns([3, 3, 2])
    eh_comemorativa = _cc1.checkbox(
        "🎉 É data comemorativa?",
        key=f"cal_comemorativa_{fk}",
        help="Marca como data festiva — exibe badge especial e balões na tela de Frequência.",
    )
    tem_aula = False
    if eh_comemorativa:
        tem_aula = _cc2.checkbox(
            "✅ Haverá aula nesta data?",
            key=f"cal_tem_aula_{fk}",
            help=(
                "Marcado → data festiva COM aula (badge + balões, chamada liberada).\n"
                "Desmarcado → feriado comemorativo SEM aula (bloqueia chamada + badge)."
            ),
        )
    else:
        _cc2.markdown(
            "<small style='color:#94A3B8;'>← Marque 'comemorativa' para habilitar</small>",
            unsafe_allow_html=True,
        )

    _cc3.markdown("<br>", unsafe_allow_html=True)
    btn_reg_cal = _cc3.button(
        "✅ REGISTRAR", type="primary", use_container_width=True, key=f"cal_btn_reg_{fk}"
    )

    if btn_reg_cal:
        with st.spinner("Salvando…"):
            ok = registrar_dia_sem_aula(
                str(novo_dia_cal),
                motivo_cal_txt,
                criado_por=st.session_state.get("usuario_logado", "sistema"),
                eh_comemorativa=eh_comemorativa,
                tem_aula=tem_aula,
            )
        if ok:
            # Invalida cache de comemorativas para o badge atualizar imediatamente
            try:
                get_datas_comemorativas_bd.clear()
            except Exception:
                pass
            if tem_aula and eh_comemorativa:
                msg = f"🎉 {novo_dia_cal.strftime('%d/%m/%Y')} registrada como data comemorativa com aula!"
            elif eh_comemorativa:
                msg = f"🎊 {novo_dia_cal.strftime('%d/%m/%Y')} registrada como feriado comemorativo (sem aula)."
            else:
                msg = f"📌 {novo_dia_cal.strftime('%d/%m/%Y')} registrado como dia sem aula."
            st.toast(msg, icon="✅")
            st.session_state["cal_form_key"] = fk + 1
            st.rerun()
        else:
            st.error("❌ Falha ao registrar.")
            with st.expander("ℹ️ SQL para criar/migrar a tabela (execute no Supabase)", expanded=True):
                st.code("""
-- Criação inicial
CREATE TABLE IF NOT EXISTS dias_sem_aula (
    id              uuid DEFAULT gen_random_uuid() PRIMARY KEY,
    data            date NOT NULL UNIQUE,
    motivo          text DEFAULT '',
    criado_em       timestamptz DEFAULT now(),
    criado_por      text DEFAULT '',
    eh_comemorativa boolean DEFAULT false,
    tem_aula        boolean DEFAULT false
);
ALTER TABLE dias_sem_aula ENABLE ROW LEVEL SECURITY;
CREATE POLICY "allow_all" ON dias_sem_aula
    FOR ALL USING (true) WITH CHECK (true);

-- Migração (se a tabela já existe — execute só as linhas ALTER)
ALTER TABLE dias_sem_aula ADD COLUMN IF NOT EXISTS eh_comemorativa boolean DEFAULT false;
ALTER TABLE dias_sem_aula ADD COLUMN IF NOT EXISTS tem_aula        boolean DEFAULT false;
                """, language="sql")

    st.markdown("---")

    # ── Lista automática ───────────────────────────────────────────────────
    st.markdown("### 📋 Registros existentes")
    fl1, fl2 = st.columns([2, 2])
    cal_ini = fl1.date_input(
        "De:", value=hoje_cal - datetime.timedelta(days=180),
        format="DD/MM/YYYY", key="cal_lista_ini"
    )
    cal_fim = fl2.date_input(
        "Até:", value=hoje_cal + datetime.timedelta(days=90),
        format="DD/MM/YYYY", key="cal_lista_fim"
    )

    # Legenda visual
    st.markdown(
        "<div style='display:flex;gap:18px;flex-wrap:wrap;margin:6px 0 14px;font-size:12px;color:#475569;'>"
        "<span>📌 Sem aula</span>"
        "<span>🎊 Comemorativa sem aula</span>"
        "<span>🎉 Comemorativa <strong>com aula</strong> (badge ativo)</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    df_cal = get_dias_sem_aula_periodo_df(str(cal_ini), str(cal_fim))
    if df_cal.empty:
        st.info("Nenhuma data especial registrada no período selecionado.")
    else:
        st.markdown(f"**{len(df_cal)} dia(s) encontrado(s):**")
        for _, row_cal in df_cal.iterrows():
            try:
                d_obj   = datetime.date.fromisoformat(str(row_cal["data"]))
                weekday = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"][d_obj.weekday()]
                d_disp  = f"{d_obj.strftime('%d/%m/%Y')} ({weekday})"
            except Exception:
                d_disp = str(row_cal["data"])

            motivo_d  = str(row_cal.get("motivo", "") or "—")
            criado_d  = str(row_cal.get("criado_por", "") or "sistema")
            _comemorativa = bool(row_cal.get("eh_comemorativa", False))
            _tem_aula     = bool(row_cal.get("tem_aula", False))

            # Ícone e cor conforme tipo
            if _comemorativa and _tem_aula:
                _ico  = "🎉"
                _cor  = "#92400E"
                _tag  = "<span style='font-size:10px;font-weight:700;background:#FEF3C7;color:#92400E;border-radius:6px;padding:1px 6px;'>COMEMORATIVA · COM AULA</span>"
            elif _comemorativa:
                _ico  = "🎊"
                _cor  = "#5B21B6"
                _tag  = "<span style='font-size:10px;font-weight:700;background:#EDE9FE;color:#5B21B6;border-radius:6px;padding:1px 6px;'>COMEMORATIVA · SEM AULA</span>"
            else:
                _ico  = "📌"
                _cor  = "#DC2626"
                _tag  = ""

            col_d, col_m, col_x = st.columns([3, 5, 1])
            col_d.markdown(
                f"<span style='color:{_cor};font-weight:700;'>{_ico} {d_disp}</span> {_tag}",
                unsafe_allow_html=True,
            )
            col_m.markdown(
                f"<small style='color:#64748B;'>{motivo_d} · <em>por {criado_d}</em></small>",
                unsafe_allow_html=True,
            )
            if col_x.button("🗑️", key=f"cal_del_{row_cal['data']}", help="Remover este dia"):
                remover_dia_sem_aula(str(row_cal["data"]))
                try:
                    get_datas_comemorativas_bd.clear()
                except Exception:
                    pass
                st.toast("🗑️ Registro removido.", icon="🗑️")
                st.rerun()


# ==============================================================================
# 🧭 NAVEGAÇÃO INTERNA E DASHBOARD
# ==============================================================================

# Executa as ferramentas de topo (Seletor de Tema e Botão do Drive) acima do menu principal
renderizar_seletor_tema()

# ── Verificação de permissão de menu ─────────────────────────────────────────
def _menu_liberado(chave: str) -> bool:
    """True se o menu está liberado para o usuário logado.
    SuperAdmin sempre tem acesso total. Default sem registro = liberado.
    """
    if st.session_state.get("perfil") == "SuperAdmin":
        return True
    uid = st.session_state.get("usuario_id", "")
    if not uid:
        return True
    perms = st.session_state.get("_menu_perms_cache")
    if perms is None:
        from database import get_menu_permissoes_usuario, get_menu_perm_version
        perms = get_menu_permissoes_usuario(uid)
        st.session_state["_menu_perms_cache"] = perms
        # Armazena a versão atual para detectar mudanças futuras
        try:
            st.session_state["_menu_perms_version"] = get_menu_perm_version(uid)
        except Exception:
            pass
    return perms.get(chave, True)


_CHAVES_MENU = {
    "Principal":       "principal",
    "Frequência":      "frequencia",
    "Portal do Aluno": "portal_aluno",
    "Relatórios & BI": "relatorios_bi",
    "Gestor":          "gestor",
}

# ── Verifica se permissões do usuário foram alteradas por outro admin ─────────
# Executa uma vez por render; se a versão no banco for diferente da cached,
# limpa o cache para que _menu_liberado() recarregue do banco neste mesmo render.
_uid_perm_check = st.session_state.get("usuario_id", "")
if (
    _uid_perm_check
    and st.session_state.get("perfil") != "SuperAdmin"
    and st.session_state.get("_menu_perms_cache") is not None
):
    try:
        from database import get_menu_perm_version as _get_perm_ver
        _db_ver = _get_perm_ver(_uid_perm_check)
        if _db_ver and _db_ver != st.session_state.get("_menu_perms_version", ""):
            st.session_state.pop("_menu_perms_cache", None)
            st.session_state["_menu_perms_version"] = _db_ver
    except Exception:
        pass

# ── Lista de páginas disponíveis (usada para lógica de redirect) ─────────────
menu = [
    m for m in ["Principal", "Frequência", "Portal do Aluno", "Relatórios & BI", "Gestor"]
    if _menu_liberado(_CHAVES_MENU[m])
]

# Se a página ativa foi revogada, redirecionar para a primeira disponível
_TOP_LEVEL_PAGES = {"Principal", "Frequência", "Portal do Aluno", "Relatórios & BI", "Gestor"}
_current_page = st.session_state.get("menu_atual", "Principal")
if _current_page in _TOP_LEVEL_PAGES and _current_page not in menu:
    if menu:
        st.session_state.menu_atual = menu[0]
    else:
        st.session_state.clear()
        st.rerun()

# ── Slugs de telemetria por destino ──────────────────────────────────────────
_NAV_SLUG = {
    "Principal":        "nav_principal",
    "Frequência":       "nav_frequencia",
    "Portal do Aluno":  "nav_portal_do_aluno",
    "Relatórios & BI":  "nav_relatorios_e_bi",
    "Gestor":           "nav_gestor",
}


def _navegar(destino: str) -> None:
    """Muda de página e registra telemetria — sem depender de key de widget."""
    st.session_state.menu_atual = destino
    _slug = _NAV_SLUG.get(destino)
    if _slug:
        try:
            from database import registrar_telemetria as _rt
            _rt(_slug)
        except Exception:
            pass
    st.rerun()


# ── CSS da sidebar de navegação ───────────────────────────────────────────────
st.markdown("""
<style>
/* ─── Sidebar: fundo e padding ──────────────────────────────────────── */
section[data-testid="stSidebar"] > div:first-child {
    padding-top: 0 !important;
    background: #FFFFFF !important;
    border-right: 1px solid #E2E8F0 !important;
}
/* ─── Botão de nav — inativo ─────────────────────────────────────────── */
section[data-testid="stSidebar"] button[kind="secondary"] {
    background: transparent !important;
    border: none !important;
    color: #374151 !important;
    font-weight: 600 !important;
    font-size: 13.5px !important;
    text-align: left !important;
    padding: 7px 12px !important;
    border-radius: 8px !important;
    margin-bottom: 2px !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background: #EFF6FF !important;
    color: #1D4ED8 !important;
}
/* ─── Botão de nav — ativo ───────────────────────────────────────────── */
section[data-testid="stSidebar"] button[kind="primary"] {
    background: #EFF6FF !important;
    color: #1D4ED8 !important;
    font-weight: 700 !important;
    font-size: 13.5px !important;
    border-radius: 8px !important;
    border: none !important;
    border-left: 3px solid #1D4ED8 !important;
    padding: 7px 12px 7px 9px !important;
    margin-bottom: 2px !important;
    box-shadow: none !important;
}
/* ─── Rótulos de grupo ───────────────────────────────────────────────── */
.sb-group-label {
    display: block;
    font-size: 10px;
    font-weight: 800;
    color: #9CA3AF;
    letter-spacing: 1.3px;
    text-transform: uppercase;
    margin: 16px 0 4px 2px;
}
.sb-hr { border: none; border-top: 1px solid #F1F5F9; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar: construção com grupos ────────────────────────────────────────────
_pagina_ativa = st.session_state.get("menu_atual", "Principal")


def _sb_btn(label: str, destino: str) -> None:
    """Botão de navegação — primary quando ativo, secondary quando inativo."""
    _ativo = _pagina_ativa == destino
    if st.sidebar.button(
        label,
        key=f"_sb_{destino.replace(' ', '_').replace('&', 'e')}",
        use_container_width=True,
        type="primary" if _ativo else "secondary",
    ):
        _navegar(destino)


with st.sidebar:
    # ── Cabeçalho: logo do sistema ────────────────────────────────────────────
    try:
        from utils.identidade import get_config as _gc_sb, get_logo_data_url as _gld_sb
        _cfg_sb   = _gc_sb()
        _logo_sb  = _gld_sb(_cfg_sb.get("logo_principal", "logo-imbra.png"))
        if _logo_sb:
            st.markdown(
                f"<div style='text-align:center;padding:14px 8px 6px;'>"
                f"<img src='{_logo_sb}' style='height:40px;object-fit:contain;'>"
                f"</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div style='text-align:center;font-size:26px;padding:12px 0 4px;'>"
                "🏃</div>",
                unsafe_allow_html=True,
            )
    except Exception:
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    # ── Cartão do operador ────────────────────────────────────────────────────
    _sb_nome = st.session_state.get("usuario_nome", "") or ""
    _sb_perf = st.session_state.get("perfil", "")
    _sb_ini  = (_sb_nome[0] if _sb_nome else "?").upper()
    st.markdown(
        f"<div style='display:flex;align-items:center;gap:8px;"
        f"padding:6px 4px 8px;'>"
        f"<div style='width:30px;height:30px;border-radius:50%;"
        f"background:linear-gradient(135deg,#3B82F6,#06B6D4);"
        f"display:flex;align-items:center;justify-content:center;"
        f"color:#fff;font-weight:900;font-size:13px;flex-shrink:0;'>"
        f"{_sb_ini}</div>"
        f"<div><p style='margin:0;font-size:12px;font-weight:700;color:#0F172A;"
        f"line-height:1.2;'>"
        f"{_sb_nome.split()[0] if _sb_nome else 'Operador'}</p>"
        f"<p style='margin:0;font-size:10px;color:#64748B;'>{_sb_perf}</p>"
        f"</div></div>",
        unsafe_allow_html=True,
    )
    st.markdown("<hr class='sb-hr'>", unsafe_allow_html=True)

    # ── 🏠 GERAL ──────────────────────────────────────────────────────────────
    st.markdown("<span class='sb-group-label'>🏠 Geral</span>",
                unsafe_allow_html=True)
    _sb_btn("🏠 Início", "Principal")

    # ── 🏃 OPERACIONAL ────────────────────────────────────────────────────────
    _op_cfg = [
        ("✅ Frequência",           "Frequência",         "frequencia"),
        ("🩺 Portal do Aluno",      "Portal do Aluno",    "portal_aluno"),
        ("💙 Radar de Inativos",    "Radar de Inativos",  "gestor_radar"),
    ]
    _op_disp = [(l, d) for l, d, c in _op_cfg if _menu_liberado(c)]
    if _op_disp:
        st.markdown("<span class='sb-group-label'>🏃 Operacional</span>",
                    unsafe_allow_html=True)
        for _lbl, _dst in _op_disp:
            _sb_btn(_lbl, _dst)

    # ── ⚙️ ADMINISTRATIVO — oculto se todas as chaves estiverem bloqueadas ───
    _adm_cfg = [
        ("📊 Relatórios & BI", "Relatórios & BI", "relatorios_bi"),
        ("🎯 Gestor",          "Gestor",           "gestor"),
    ]
    _adm_disp = [(l, d) for l, d, c in _adm_cfg if _menu_liberado(c)]
    if _adm_disp:
        st.markdown("<span class='sb-group-label'>⚙️ Administrativo</span>",
                    unsafe_allow_html=True)
        for _lbl, _dst in _adm_disp:
            _sb_btn(_lbl, _dst)

    # ── Sair ──────────────────────────────────────────────────────────────────
    st.markdown("<hr class='sb-hr'>", unsafe_allow_html=True)
    if st.sidebar.button("🔓 Sair", key="_sb_sair", use_container_width=True,
                         type="secondary"):
        st.session_state.clear()
        st.rerun()

# --- DASHBOARD PRINCIPAL ---
if st.session_state.menu_atual == "Principal":
    # ── CSS Global do Dashboard ─────────────────────────────────────────────
    st.markdown("""
<style>
.hg-avatar-wrap {
  display: flex;
  align-items: center;
  justify-content: center;
}
.hg-avatar {
  display: block;
  width: 42px !important;
  height: 42px !important;
  min-width: 42px;
  min-height: 42px;
  max-width: 42px;
  max-height: 42px;
  aspect-ratio: 1 / 1;
  border-radius: 50%;
  object-fit: cover;
  object-position: center center;
  flex-shrink: 0;
  box-shadow: 0 0 0 2.5px #3B82F6, 0 2px 8px rgba(59,130,246,0.25);
  transition: transform 0.25s cubic-bezier(0.34,1.56,0.64,1), box-shadow 0.25s ease;
  cursor: zoom-in;
}
.hg-avatar:hover {
  transform: scale(3.5);
  box-shadow: 0 0 0 2.5px #3B82F6, 0 12px 36px rgba(0,0,0,0.5);
  z-index: 9999;
  position: relative;
}
.hg-initials {
  display: flex;
  align-items: center;
  justify-content: center;
  width: 42px;
  height: 42px;
  min-width: 42px;
  min-height: 42px;
  aspect-ratio: 1 / 1;
  border-radius: 50%;
  flex-shrink: 0;
  background: linear-gradient(135deg, #3B82F6, #06B6D4);
  color: #fff;
  font-weight: 900;
  font-size: 14px;
  box-shadow: 0 0 0 2.5px #3B82F6, 0 2px 8px rgba(59,130,246,0.2);
}
.hg-niver-hoje  { color: #10B981; font-weight: 800; }
.hg-niver-breve { color: #F59E0B; font-weight: 700; }
.hg-niver-passou{ color: #94A3B8; }
</style>""", unsafe_allow_html=True)

    # ── Cabeçalho: Saudação + Data + Relógio SP ─────────────────────────────
    _hoje = datetime.date.today()
    _ds = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
           "Sexta-feira", "Sábado", "Domingo"][_hoje.weekday()]
    _meses_pt_full = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                      "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    _meses_abrev   = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
    _nome_curto = (
        st.session_state.usuario_nome.split()[0]
        if st.session_state.usuario_nome else "Gestor"
    )

    _col_greet, _col_clock = st.columns([3, 1], gap="small", vertical_alignment="center")

    with _col_greet:
        st.markdown(
            f"<div style='margin-bottom:4px;'>"
            f"<h3 style='color:#0A2540;margin:0;font-size:1.2rem;font-weight:800;line-height:1.3;'>"
            f"Olá, {_nome_curto} 👋</h3>"
            f"<p style='color:#475569;font-size:13px;font-weight:600;margin:2px 0 0;'>"
            f"📅 {_ds}, {_hoje.day:02d} de {_meses_pt_full[_hoje.month-1]} de {_hoje.year}"
            f"&nbsp;&nbsp;"
            f"<span style='background:#EFF6FF;border:1px solid #BFDBFE;border-radius:6px;"
            f"padding:2px 8px;font-size:11px;font-weight:700;color:#1E40AF;'>"
            f"● {st.session_state.perfil}</span>"
            f"</p></div>",
            unsafe_allow_html=True,
        )

    with _col_clock:
        import streamlit.components.v1 as _stc
        _stc.html(
            """
<div style='text-align:right; padding:4px 0; font-family:sans-serif;'>
  <div id="ck"
       style='font-size:1.65rem; font-weight:900; color:#1E40AF;
              font-family:monospace; letter-spacing:2px; line-height:1;'>
    --:--:--
  </div>
  <div style='font-size:10px; color:#94A3B8; font-weight:600; margin-top:2px;'>
    🕒 Horário de São Paulo
  </div>
</div>
<script>
(function(){
  function tick(){
    var el = document.getElementById('ck');
    if(!el){ setTimeout(tick,150); return; }
    el.textContent = new Date().toLocaleTimeString('pt-BR',
      {timeZone:'America/Sao_Paulo',hour:'2-digit',minute:'2-digit',second:'2-digit'});
  }
  tick();
  setInterval(tick, 1000);
})();
</script>""",
            height=58,
            scrolling=False,
        )

    st.markdown("<hr style='margin:8px 0 10px;border-color:#E2E8F0;'/>", unsafe_allow_html=True)

    # ── Dialog: Agendar Nova Avaliação ───────────────────────────────────────
    @st.dialog("✏️ Gerenciar Agendamento")
    def _dialog_gerenciar_ag(ag_id, nome, data_atual, hora_atual):
        import datetime as _dt3
        from database import (
            concluir_ou_cancelar_agendamento as _cca,
            atualizar_agendamento as _uag,
        )

        st.markdown(
            f"<div style='font-size:13px;color:#0A2540;font-weight:700;"
            f"margin-bottom:12px;'>👤 {nome}</div>",
            unsafe_allow_html=True,
        )

        # ── Editar data / hora ──────────────────────────────────────────
        try:
            _data_pre = _dt3.date.fromisoformat(str(data_atual)[:10])
        except Exception:
            _data_pre = _dt3.date.today() + _dt3.timedelta(days=1)
        try:
            _h, _m = str(hora_atual)[:5].split(":")
            _hora_pre = _dt3.time(int(_h), int(_m))
        except Exception:
            _hora_pre = _dt3.time(11, 0)

        _eg1, _eg2 = st.columns(2)
        _nova_data = _eg1.date_input(
            "📅 Data:", value=_data_pre,
            min_value=_dt3.date.today(),
            key="gag_data", format="DD/MM/YYYY",
        )
        _nova_hora = _eg2.time_input(
            "🕐 Horário:", value=_hora_pre,
            key="gag_hora", step=1800,
        )

        if st.button("💾 Salvar Alteração", type="primary", use_container_width=True):
            _ok_u, _msg_u = _uag(ag_id, str(_nova_data), str(_nova_hora)[:5])
            if _ok_u:
                get_agendamentos_pendentes.clear()
                st.rerun()
            else:
                st.error(f"Erro: {_msg_u}")

        st.divider()

        # ── Exclusão com confirmação inline ────────────────────────────
        _key_conf = f"gag_conf_{ag_id}"
        if not st.session_state.get(_key_conf):
            if st.button("🗑️ Excluir Agendamento", use_container_width=True):
                st.session_state[_key_conf] = True
                st.rerun()
        else:
            st.warning(f"Confirma excluir o agendamento de **{nome}**?")
            _cd1, _cd2 = st.columns(2)
            if _cd1.button("✅ Sim, excluir", type="primary", use_container_width=True):
                _ok_x, _msg_x = _cca(ag_id, "Cancelado")
                if _ok_x:
                    st.session_state.pop(_key_conf, None)
                    get_agendamentos_pendentes.clear()
                    st.rerun()
                else:
                    st.error(f"Erro: {_msg_x}")
            if _cd2.button("↩️ Não, voltar", use_container_width=True):
                st.session_state.pop(_key_conf, None)
                st.rerun()

    @st.dialog("📅 Agendar Nova Avaliação")
    def _dialog_nova_aval():
        import datetime as _dt2
        from database import buscar_alunos_geral as _bag, criar_agendamento as _cag
        from utils.busca_aluno import busca_aluno_widget as _baw, filtrar_alunos_df as _faf
        st.caption("Busque o aluno, confirme a data e o horário.")
        _busca_d = _baw("dag_busca", placeholder="🔍 Digite pelo menos 3 letras do nome...")
        _aluno_d = None
        if _busca_d and len(_busca_d.strip()) >= 3:
            _df_d = _faf(_bag(""), _busca_d, cols=["nome", "turma"])
            if _df_d.empty:
                st.warning("Nenhum aluno encontrado.")
            else:
                _map_d = {
                    f"{r['nome']}  ({str(r.get('turma',''))[:18]})": r.to_dict()
                    for _, r in _df_d.head(10).iterrows()
                }
                _k_d = st.selectbox("Selecionar:", list(_map_d.keys()), key="dag_sel",
                                    label_visibility="collapsed")
                _aluno_d = _map_d[_k_d]
        # Data — próximo dia útil como padrão
        _hoje_d = _dt2.date.today()
        _prox_d = _hoje_d + _dt2.timedelta(days=1)
        while _prox_d.weekday() >= 5:
            _prox_d += _dt2.timedelta(days=1)
        _c1d, _c2d = st.columns(2)
        _data_d = _c1d.date_input("📅 Data:", value=_prox_d, min_value=_hoje_d,
                                   key="dag_data", format="DD/MM/YYYY")
        _hora_d = _c2d.time_input("🕐 Horário:", value=_dt2.time(11, 0), key="dag_hora",
                                   step=1800)
        st.divider()
        if st.button("✅ Confirmar Agendamento", type="primary",
                     use_container_width=True, disabled=(_aluno_d is None)):
            _ok_d, _msg_d = _cag(
                _aluno_d["id"], str(_data_d), str(_hora_d)[:5], "Avaliação"
            )
            if _ok_d:
                get_agendamentos_pendentes.clear()
                st.rerun()
            else:
                st.error(f"Erro: {_msg_d}")

    # ════════════════════════════════════════════════════════════════════════
    # CARD HEROICO — Atalho Rápido para a Chamada do Dia
    # ════════════════════════════════════════════════════════════════════════
    if _menu_liberado("frequencia") and _menu_liberado("freq_chamada_tablet"):
        with st.container(border=True):
            _hc_icon, _hc_txt, _hc_btn = st.columns(
                [0.6, 4, 2], vertical_alignment="center", gap="medium"
            )
            _hc_icon.markdown(
                "<div style='font-size:2.8rem;text-align:center;line-height:1;"
                "padding:4px 0;'>📋</div>",
                unsafe_allow_html=True,
            )
            with _hc_txt:
                st.markdown(
                    "<p style='margin:0;font-size:1.05rem;font-weight:800;"
                    "color:#0A2540;line-height:1.3;'>Iniciar Chamada do Dia</p>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f"<p style='margin:2px 0 0;font-size:0.82rem;color:#64748B;"
                    f"font-weight:600;'>📅 {_ds}, "
                    f"{_hoje.day:02d} de {_meses_pt_full[_hoje.month-1]} "
                    f"de {_hoje.year}</p>",
                    unsafe_allow_html=True,
                )
            with _hc_btn:
                if st.button(
                    "🚀 Fazer Chamada",
                    key="hg_hero_chamada",
                    type="primary",
                    use_container_width=True,
                ):
                    st.session_state["_freq_ir_tablet"] = True
                    _navegar("Frequência")

    st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

    # ── Layout principal: Grid (esq) + Avaliações (dir) ─────────────────────
    _col_grid, _col_agenda = st.columns([3, 1], gap="large")

    # ════════════════════════════════════════════════
    # COLUNA DIREITA — Próximas Avaliações (sempre visível)
    # ════════════════════════════════════════════════
    with _col_agenda:
        # CSS: reduz fonte do nome nos cards de agendamento em 35%
        st.markdown("""<style>
        div[data-testid="column"] + div[data-testid="column"] button p,
        div[data-testid="column"] + div[data-testid="column"] button span {
            font-size: 9px !important;
            line-height: 1.3 !important;
        }
        div[data-testid="column"] + div[data-testid="column"]
          div[data-testid="stVerticalBlockBorderWrapper"] button {
            padding: 3px 6px !important;
            min-height: unset !important;
        }
        </style>""", unsafe_allow_html=True)

        _ag_hd_col, _ag_add_col = st.columns([3, 1], vertical_alignment="center")
        _ag_hd_col.markdown(
            "<p style='font-weight:800;color:#0A2540;font-size:0.95rem;margin:0 0 6px;'>"
            "🗓️ Próximas Avaliações</p>",
            unsafe_allow_html=True,
        )
        with _ag_add_col:
            if st.button("➕", key="hg_nova_aval", help="Agendar Nova Avaliação",
                         use_container_width=True):
                _dialog_nova_aval()

        _agendamentos = get_agendamentos_pendentes(limite=8)
        if _agendamentos:
            for _ag in _agendamentos:
                _aluno_data = _ag.get("alunos") or {}
                _nm  = (_aluno_data.get("nome") or _ag.get("nome_aluno") or "—")
                _aid = _aluno_data.get("id") or _ag.get("aluno_id")
                _hr  = (_ag.get("horario") or "—")
                _dt  = str(_ag.get("data_agendamento") or _ag.get("data") or "")
                _dt_fmt = ""
                if _dt and len(_dt) >= 10:
                    try:
                        _dp = datetime.date.fromisoformat(_dt[:10])
                        _dias_sem = ["seg","ter","qua","qui","sex","sáb","dom"]
                        _dia_sem = _dias_sem[_dp.weekday()]
                        _dt_fmt = f"{_dia_sem} {_dp.day:02d}/{_dp.month:02d}"
                    except Exception:
                        _dt_fmt = _dt[:5]
                _nm_curto = (_nm[:16] + "…") if len(_nm) > 16 else _nm
                with st.container(border=True):
                    st.markdown(
                        f"<div style='font-size:11px;color:#64748B;margin-bottom:2px;'>"
                        f"🕒 <b>{_hr}</b>{(' · ' + _dt_fmt) if _dt_fmt else ''}</div>",
                        unsafe_allow_html=True,
                    )
                    _ag_id_str = _ag.get("id", "")
                    _c_nm, _c_del = st.columns([4, 1], vertical_alignment="center")
                    if _aid:
                        if _c_nm.button(
                            _nm_curto,
                            key=f"hg_ag_{_ag_id_str}",
                            use_container_width=True,
                            help=f"Abrir prontuário: {_nm} → Nova Medição",
                        ):
                            from database import buscar_aluno_por_id
                            _al = buscar_aluno_por_id(_aid) or _aluno_data
                            st.session_state.aluno_prontuario = _al
                            st.session_state.origem_prontuario = "Principal"
                            st.session_state.prontuario_aba = "Nova Medição"
                            st.session_state.menu_atual = "Portal do Aluno"
                            st.rerun()
                    else:
                        _c_nm.markdown(
                            f"<div style='font-size:8px;font-weight:600;"
                            f"color:#0F172A;'>{_nm_curto}</div>",
                            unsafe_allow_html=True,
                        )
                    if _ag_id_str and _c_del.button(
                        "✏️", key=f"hg_ag_del_{_ag_id_str}",
                        help="Editar ou excluir este agendamento",
                    ):
                        _dialog_gerenciar_ag(_ag_id_str, _nm, _dt, _hr)
        else:
            st.info("Agenda livre ✅", icon=None)

    # ════════════════════════════════════════════════
    # COLUNA ESQUERDA — Grid de Alunos
    # ════════════════════════════════════════════════
    with _col_grid:
        from database import (
            buscar_alunos_geral,
            get_snapshot_home_grid as _get_snap_home,
            salvar_snapshot_home_grid as _save_snap_home,
            computar_snapshot_home_grid as _computar_snap_home,
        )

        # ── Catálogo de colunas toggleáveis (chave, rótulo UI, largura tela) ──
        _HG_COL_DEFS = [
            ("turma",    "Turma",        0.9),
            ("aniv",     "Aniversário",  0.8),
            ("freq",     "Freq.+Pres.",  1.6),
            ("ausencia", "⚠️ Ausência",   1.0),
            ("atestado", "Atestado",     1.0),
            ("pa",       "Última PA",    1.0),
            ("anam",     "Anamnese",     1.0),
            ("wap",      "WhatsApp",     0.8),
        ]

        # ── Linha de título + busca + turma + sexo + colunas + Processar em Lote ──
        _hg_c_titulo, _hg_c_busca, _hg_c_turma, _hg_c_sexo, _hg_c_cols, _hg_c_lote = st.columns(
            [1.4, 2.5, 1.8, 1.4, 0.75, 1.3], vertical_alignment="bottom", gap="small"
        )
        _hg_c_titulo.markdown(
            "<p style='font-weight:800;color:#0A2540;font-size:1.05rem;margin:0;'>👥 Alunos Ativos</p>",
            unsafe_allow_html=True,
        )
        from utils.busca_aluno import busca_aluno_widget as _baw_hg, filtrar_alunos_df as _faf_hg
        _hg_busca = _baw_hg(
            "hg_busca",
            container=_hg_c_busca,
            placeholder="🔍 Nome ou turma…",
        )

        # ── Snapshot: leitura + botão Processar em Lote ──────────────────────
        _snap_home = _get_snap_home()
        _snap_ts   = _snap_home.get("gerado_em", "")

        with _hg_c_lote:
            if st.button(
                "⚙️ Processar em Lote",
                key="hg_proc_lote",
                use_container_width=True,
                help=(
                    "Recalcula presenças, atestados, PA e anamnese de todos os alunos "
                    "e salva o resultado para carregamento instantâneo."
                ),
            ):
                with st.spinner("⏳ Calculando painel… aguarde."):
                    _snap_new = _computar_snap_home()
                    _ok_snap, _msg_snap = _save_snap_home(_snap_new)
                    # Limpar caches in-memory para que leituras futuras usem o snapshot
                    load_frequencia_ultima_presenca.clear()
                    load_atestados_vencimento.clear()
                    load_total_presencas_todos.clear()
                    _snap_home = _snap_new
                    _snap_ts   = _snap_home.get("gerado_em", "")
                if _ok_snap:
                    st.toast("✅ Painel atualizado com sucesso!", icon="✅")
                    st.rerun()
                else:
                    st.error(f"Erro ao salvar painel: {_msg_snap}")

        # Popover de visibilidade de colunas
        with _hg_c_cols:
            with st.popover("⊞", use_container_width=True, help="Mostrar/ocultar colunas"):
                st.caption("**Colunas visíveis:**")
                for _ck, _cl, _ in _HG_COL_DEFS:
                    st.checkbox(_cl, value=True, key=f"hg_col_{_ck}")

        if _snap_ts:
            st.caption(f"📦 Painel processado em: **{_snap_ts}**")
        else:
            st.info(
                "⚠️ Painel ainda não processado. Clique em **⚙️ Processar em Lote** "
                "para gerar os dados e acelerar a abertura do sistema.",
                icon=None,
            )

        # ── Checkbox: grid desabilitado por padrão para evitar crash na abertura ──
        _grid_habilitado = st.checkbox(
            "📊 Exibir painel de alunos",
            value=False,
            key="hg_grid_ativo",
            help=(
                "Desabilitado por padrão para o sistema abrir rápido e sem travamentos. "
                "Ative aqui quando precisar consultar o painel completo de alunos."
            ),
        )
        if not _grid_habilitado:
            st.info(
                "💡 Painel desabilitado. Marque **Exibir painel de alunos** acima para carregar os dados.",
                icon=None,
            )

        # ── Carregar dados — somente se o grid estiver habilitado ────────────
        if _grid_habilitado:
            _df_alunos = buscar_alunos_geral("")

            if _snap_ts:
                _recs_up = _snap_home.get("ultima_presenca_recs", [])
                _df_ultima = pd.DataFrame(_recs_up) if _recs_up else pd.DataFrame(columns=["id", "ultima_presenca"])

                _recs_at = _snap_home.get("atestados_recs", [])
                _df_atestad = pd.DataFrame(_recs_at) if _recs_at else pd.DataFrame(columns=["id", "data_vencimento_atestado"])

                _recs_tp = _snap_home.get("total_presencas_recs", [])
                if _recs_tp:
                    _df_total_pres = pd.DataFrame(_recs_tp)
                else:
                    # Snapshot não tem presencas — busca ao vivo (evita zerar a coluna)
                    _df_total_pres = load_total_presencas_todos()
            else:
                _df_ultima     = load_frequencia_ultima_presenca()
                _df_atestad    = load_atestados_vencimento()
                _df_total_pres = load_total_presencas_todos()
        else:
            _df_alunos = pd.DataFrame()

        # Templates WhatsApp — carregados uma vez, fora do loop.
        # Cada gatilho da tabela crm_templates (painel Configurações > Mensagens)
        # é mapeado com exatidão: niver_hoje (Dia Exato) / niver_passou (Atrasado) /
        # evasao_nunca / evasao_60 / evasao_80 / assiduo_top.
        # "Aviso Prévio" (niver_futuro) foi descontinuado: futuros não recebem parabéns.
        try:
            from utils.niver_automatico import montar_link_whatsapp, personalizar_mensagem, montar_mensagem_niver, get_parabenizados_dict
            from database import get_crm_templates
            _df_tpl = get_crm_templates()
            # Dicionário {gatilho: mensagem} — fonte única e auditada
            _tpl_map: dict = {}
            if not _df_tpl.empty and "gatilho" in _df_tpl.columns:
                for _, _tr in _df_tpl.iterrows():
                    _tpl_map[str(_tr.get("gatilho", ""))] = str(_tr.get("mensagem", ""))
            _parab_dict = get_parabenizados_dict()
            _wapp_ok = True
        except Exception:
            _tpl_map = {}
            _parab_dict = {}
            _wapp_ok = False

        if not _df_alunos.empty:
            # Merge com última presença
            if not _df_ultima.empty:
                _df_hg = _df_alunos.merge(_df_ultima, on="id", how="left")
            else:
                _df_hg = _df_alunos.copy()
                _df_hg["ultima_presenca"] = pd.NaT

            # Merge com vencimento de atestado
            if not _df_atestad.empty:
                _df_hg = _df_hg.merge(_df_atestad, on="id", how="left")
            else:
                _df_hg["data_vencimento_atestado"] = pd.NaT

            _df_hg["ultima_presenca"] = pd.to_datetime(_df_hg["ultima_presenca"], errors="coerce")

            # Merge com total histórico de presenças
            if not _df_total_pres.empty:
                _df_hg = _df_hg.merge(_df_total_pres, on="id", how="left")
            else:
                _df_hg["total_presencas_hist"] = 0
            _df_hg["total_presencas_hist"] = _df_hg["total_presencas_hist"].fillna(0).astype(int)

            # Merge com última PA
            try:
                from database import get_ultima_pa_todos
                from views.prontuario_dashboard import _pa_compact_html
                if _snap_ts and _snap_home.get("ultima_pa"):
                    _snap_pa_raw = _snap_home["ultima_pa"]
                    _pa_dict = {}
                    for _sk, _sv in _snap_pa_raw.items():
                        _pa_dict[_sk] = _sv
                        if _sk.isdigit():
                            _pa_dict[int(_sk)] = _sv
                else:
                    _pa_dict = get_ultima_pa_todos()
                def _pa_html_row(aluno_id):
                    p = _pa_dict.get(str(aluno_id)) or _pa_dict.get(int(aluno_id) if str(aluno_id).isdigit() else aluno_id)
                    if not p:
                        return None, "—", "<span style='color:#CBD5E1;font-size:11px;'>—</span>", None, ""
                    _sis = p.get("sis")
                    _dia = p.get("dia")
                    _txt = f"{_sis}/{_dia}" if _sis and _dia else (str(_sis) if _sis else "—")
                    return (
                        _sis, _txt,
                        _pa_compact_html(_sis, _dia, p.get("pul"), p.get("cls", "normal")),
                        _dia, p.get("cls", ""),
                    )
                _df_hg[["_pa_sis", "_pa_txt", "_pa_html", "_pa_dia", "_pa_cls"]] = _df_hg["id"].apply(
                    lambda _aid: pd.Series(_pa_html_row(_aid))
                )
            except Exception:
                _df_hg["_pa_sis"]  = None
                _df_hg["_pa_txt"]  = "—"
                _df_hg["_pa_html"] = "<span style='color:#CBD5E1;font-size:11px;'>—</span>"
                _df_hg["_pa_dia"]  = None
                _df_hg["_pa_cls"]  = ""

            # Merge com status da Anamnese (avaliação clínica / prontuario_avaliacoes)
            try:
                from database import get_ultima_avaliacao_todos
                from modulos_frequencia.tab_admin import get_dias_validade_anamnese
                if _snap_ts and _snap_home.get("ultima_aval"):
                    _anam_dict = _snap_home["ultima_aval"]
                else:
                    _anam_dict = get_ultima_avaliacao_todos()
                _dias_val_anam = get_dias_validade_anamnese()
                def _anam_status_row(aluno_id):
                    _dt_str = _anam_dict.get(str(aluno_id))
                    if not _dt_str:
                        return None, "nunca"
                    try:
                        _dt_anam = pd.Timestamp(_dt_str).date()
                    except Exception:
                        return None, "nunca"
                    _dias_desde = (datetime.date.today() - _dt_anam).days
                    if _dias_desde > _dias_val_anam:
                        _st_anam = "vencida"
                    elif _dias_desde >= (_dias_val_anam - 30):
                        _st_anam = "a_vencer"
                    else:
                        _st_anam = "ok"
                    return _dt_anam, _st_anam
                _df_hg[["_anam_data", "_anam_status"]] = _df_hg["id"].apply(
                    lambda _aid: pd.Series(_anam_status_row(_aid))
                )
            except Exception:
                _df_hg["_anam_data"]   = None
                _df_hg["_anam_status"] = "nunca"

            # Birthday cols
            _df_hg["_dt_nasc"] = pd.to_datetime(_df_hg["data_nascimento"], errors="coerce")
            _df_hg["_dia_n"]   = _df_hg["_dt_nasc"].dt.day
            _df_hg["_mes_n"]   = _df_hg["_dt_nasc"].dt.month

            # Filtro de turma
            _turmas_disp = sorted(
                [t for t in _df_hg["turma"].dropna().unique().tolist() if str(t).strip()],
                key=str
            )
            _hg_turma = _hg_c_turma.selectbox(
                "Turma:", ["Todas"] + _turmas_disp,
                label_visibility="collapsed", key="hg_turma"
            )

            # Filtro de sexo — segmented_control (pills, 1 clique)
            with _hg_c_sexo:
                _hg_sexo = st.segmented_control(
                    "Sexo",
                    options=["Todos", "♀ Fem.", "♂ Masc."],
                    default="Todos",
                    key="hg_sexo",
                    label_visibility="collapsed",
                )
            _hg_sexo = _hg_sexo or "Todos"

            # Visibilidade de colunas (lida do session_state, padrão True)
            _hg_vis = {k: st.session_state.get(f"hg_col_{k}", True)
                       for k, _, _ in _HG_COL_DEFS}

            # Aplicar filtros
            _df_grid = _df_hg.copy()
            _df_grid = _faf_hg(_df_grid, _hg_busca, cols=["nome", "turma"], min_len=3)
            if _hg_turma != "Todas":
                _df_grid = _df_grid[_df_grid["turma"] == _hg_turma]
            if _hg_sexo != "Todos" and "sexo" in _df_grid.columns:
                _sexo_val = "F" if "Fem" in _hg_sexo else "M"
                # Aceita tanto "M"/"F" quanto "Masculino"/"Feminino"
                _df_grid = _df_grid[
                    _df_grid["sexo"].astype(str).str.upper().str.startswith(_sexo_val)
                ]

            # Pré-carrega vínculos estruturados de voluntariado (1 query, sem N+1)
            try:
                from database import get_acoes_todos_voluntarios as _get_acoes_vol
                _acoes_vol_dict = _get_acoes_vol()
            except Exception:
                _acoes_vol_dict = {}

            # ── Filtro de voluntários ─────────────────────────────────────────
            _hg_vol_filter = st.checkbox(
                "🤝 Somente voluntários",
                key="hg_vol_filter",
                help="Exibe apenas alunos que declararam interesse em trabalho voluntário",
            )
            _hg_acao_filter: list = []
            if _hg_vol_filter:
                _vol_col = "trabalho_voluntario_interesse"
                if _vol_col in _df_grid.columns:
                    _df_grid = _df_grid[
                        _df_grid[_vol_col].fillna("").str.strip().str.lower() == "sim"
                    ].reset_index(drop=True)

                # Multiselect de ações — opções derivadas de todos os vínculos cadastrados
                _opcoes_acoes = sorted({
                    acao
                    for acoes in _acoes_vol_dict.values()
                    for acao in acoes
                })
                if _opcoes_acoes:
                    _hg_acao_filter = st.multiselect(
                        "🔍 Filtrar por ação voluntariada:",
                        options=_opcoes_acoes,
                        default=[],
                        key="hg_acao_filter",
                        placeholder="Todas as ações (sem filtro)",
                        help="Mostra apenas voluntários vinculados às ações selecionadas",
                    )
                    if _hg_acao_filter:
                        _ids_com_acao = {
                            aid for aid, acoes in _acoes_vol_dict.items()
                            if any(a in _hg_acao_filter for a in acoes)
                        }
                        _df_grid = _df_grid[
                            _df_grid["id"].astype(str).isin(_ids_com_acao)
                        ].reset_index(drop=True)

                _n_vol = len(_df_grid)
                _caption_vol = f"🤝 **{_n_vol}** voluntário(s)"
                if _hg_acao_filter:
                    _caption_vol += f" · ação: {', '.join(_hg_acao_filter)}"
                else:
                    _caption_vol += " identificado(s) na base ativa"
                st.caption(_caption_vol)

            # ── Ao trocar de turma, limpar filtro de faixa automaticamente ──────────
            _turma_prev_ev = st.session_state.get("hg_turma_prev_ev")
            if (_turma_prev_ev is not None and
                    _turma_prev_ev != _hg_turma and
                    st.session_state.get("hg_filtro_evasao")):
                st.session_state.pop("hg_filtro_evasao", None)
                st.session_state["hg_turma_prev_ev"] = _hg_turma
                st.session_state.hg_pg = 1
                st.rerun()
            st.session_state["hg_turma_prev_ev"] = _hg_turma

            # ── Filtro por faixa de evasão (aplicado após todos os outros filtros) ─
            _hg_faixa_ev = st.session_state.get("hg_filtro_evasao")  # verde/amarelo/laranja/vermelho ou None
            if _hg_faixa_ev and "ultima_presenca" in _df_grid.columns:
                _hoje_filt = datetime.date.today()
                def _classifica_faixa(val):
                    if pd.isna(val) or val is None:
                        return "vermelho"
                    try:
                        d = (_hoje_filt - pd.Timestamp(val).date()).days
                        if d <= 7:
                            return "verde"
                        elif d <= 30:
                            return "amarelo"
                        elif d <= 60:
                            return "laranja"
                        else:
                            return "vermelho"
                    except Exception:
                        return "vermelho"
                _df_grid = _df_grid[
                    _df_grid["ultima_presenca"].map(_classifica_faixa) == _hg_faixa_ev
                ].reset_index(drop=True)

            # ── Pill de filtro de faixa ativo ─────────────────────────────────
            if _hg_faixa_ev:
                _faixa_pill_map = {
                    "verde":    ("🟢", "Ausência ≤ 7 dias",   "#D1FAE5", "#065F46"),
                    "amarelo":  ("🟡", "Ausência 8–30 dias",  "#FEF3C7", "#92400E"),
                    "laranja":  ("🟠", "Ausência 31–60 dias", "#FFEDD5", "#9A3412"),
                    "vermelho": ("🔴", "Ausência > 60 dias",  "#FEE2E2", "#991B1B"),
                }
                _pill_emoji, _pill_desc, _pill_bg, _pill_fg = _faixa_pill_map.get(
                    _hg_faixa_ev, ("🔵", _hg_faixa_ev, "#DBEAFE", "#1E3A8A")
                )
                _pill_col, _pill_btn_col = st.columns([5, 1], gap="small")
                _pill_col.markdown(
                    f"<div style='background:{_pill_bg};color:{_pill_fg};"
                    f"padding:5px 12px;border-radius:9999px;display:inline-flex;"
                    f"align-items:center;gap:6px;font-size:12px;font-weight:600;"
                    f"border:1px solid {_pill_fg}44;margin:2px 0;'>"
                    f"{_pill_emoji}&nbsp;<b>Filtro de faixa:</b>&nbsp;{_pill_desc}"
                    f"&nbsp;<span style='opacity:0.6;font-size:10px;'>"
                    f"· {len(_df_grid)} aluno(s)</span></div>",
                    unsafe_allow_html=True,
                )
                if _pill_btn_col.button(
                    "✖ Limpar filtro",
                    key="hg_pill_clear",
                    use_container_width=True,
                    help="Remover filtro de faixa de ausência e voltar ao grid completo",
                ):
                    st.session_state.pop("hg_filtro_evasao", None)
                    st.session_state.hg_pg = 1
                    st.rerun()

            # Ordenação dinâmica
            if "hg_sort_col" not in st.session_state:
                st.session_state.hg_sort_col = "nome"
                st.session_state.hg_sort_asc = True
            _sc = st.session_state.get("hg_sort_col", "nome")
            _sa = st.session_state.get("hg_sort_asc", True)
            if _sc == "aniversario":
                _df_grid = _df_grid.sort_values(
                    ["_mes_n", "_dia_n"], ascending=_sa, na_position="last"
                )
            elif _sc == "ultima_presenca":
                # Ausente há mais tempo primeiro: NaT (nunca veio) sobe ao topo
                _df_grid = _df_grid.sort_values(
                    _sc, ascending=_sa, na_position="first" if _sa else "last"
                )
            elif _sc in _df_grid.columns:
                _df_grid = _df_grid.sort_values(_sc, ascending=_sa, na_position="last")
            else:
                _df_grid = _df_grid.sort_values("nome")
            _df_grid = _df_grid.reset_index(drop=True)

            # Paginação com botões ◀ ▶
            _hg_ppp   = 20
            _hg_total = max(1, math.ceil(len(_df_grid) / _hg_ppp))
            if "hg_pg" not in st.session_state:
                st.session_state.hg_pg = 1
            # Reset para pág 1 quando filtro muda
            if st.session_state.get("hg_busca_prev") != _hg_busca or \
               st.session_state.get("hg_turma_prev") != _hg_turma:
                st.session_state.hg_pg = 1
                st.session_state.hg_busca_prev = _hg_busca
                st.session_state.hg_turma_prev = _hg_turma
            st.session_state.hg_pg = max(1, min(st.session_state.hg_pg, _hg_total))
            _hg_pg  = st.session_state.hg_pg
            _hg_ini = (_hg_pg - 1) * _hg_ppp
            _df_pag = _df_grid.iloc[_hg_ini: _hg_ini + _hg_ppp]

            # Gerador de PDF da lista home
            def _gerar_pdf_hg(df: pd.DataFrame, vis: dict | None = None, filtros: str = "") -> bytes:
                import io as _io
                import requests as _req
                from fpdf import FPDF
                from concurrent.futures import ThreadPoolExecutor
                try:
                    from PIL import Image as _Img, ImageDraw as _IDraw
                    _PIL_OK = True
                except Exception:
                    _PIL_OK = False

                df = df.reset_index(drop=True)

                # ── Sanitizador Latin-1 ──────────────────────────────────
                def _s(t):
                    return (
                        str(t)
                        .replace("\u2014", "-").replace("\u2013", "-")
                        .replace("\u2764", "").replace("\u00b7", ".")
                        .replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')
                        .encode("latin-1", errors="replace").decode("latin-1")
                    )

                # ── Download + crop circular ─────────────────────────────
                def _baixar_foto_circular(url: str):
                    if not _PIL_OK or not url or not url.startswith("http"):
                        return None
                    try:
                        resp = _req.get(url, timeout=2)
                        if resp.status_code != 200:
                            return None
                        img = _Img.open(_io.BytesIO(resp.content)).convert("RGBA")
                        w, h = img.size
                        m = min(w, h)
                        img = img.crop(((w - m) // 2, (h - m) // 2,
                                        (w + m) // 2, (h + m) // 2))
                        img = img.resize((80, 80), _Img.LANCZOS)
                        mask = _Img.new("L", (80, 80), 0)
                        _IDraw.Draw(mask).ellipse((0, 0, 80, 80), fill=255)
                        bg = _Img.new("RGBA", (80, 80), (255, 255, 255, 255))
                        bg.paste(img, mask=mask)
                        out = _io.BytesIO()
                        bg.convert("RGB").save(out, format="JPEG", quality=75)
                        out.seek(0)
                        return out
                    except Exception:
                        return None

                # ── Downloads em paralelo ────────────────────────────────
                _foto_urls = [str(a.get("foto_url") or "").strip()
                              for _, a in df.iterrows()]
                _foto_cache: dict = {}

                def _fetch_one(idx_url):
                    idx, url = idx_url
                    return idx, _baixar_foto_circular(url)

                try:
                    with ThreadPoolExecutor(max_workers=12) as _ex:
                        for _idx, _ibuf in _ex.map(_fetch_one, enumerate(_foto_urls)):
                            _foto_cache[_idx] = _ibuf
                except Exception:
                    pass

                # ── Dimensões ────────────────────────────────────────────
                _ROW_H    = 12     # altura de cada linha (mm)
                _FOTO_W   = 12     # largura da coluna foto
                _FOTO_SZ  = 9.5   # tamanho da imagem circular (mm)
                _FOTO_PX  = (_FOTO_W - _FOTO_SZ) / 2   # padding horizontal
                _FOTO_PY  = (_ROW_H - _FOTO_SZ) / 2    # padding vertical

                _tot_pdf = len(df)

                # ── Cabeçalho / Rodapé ───────────────────────────────────
                class _PDFHG(FPDF):
                    def header(self):
                        self.set_font("Helvetica", "B", 13)
                        self.cell(0, 8, _s("IMBRA - Alunos Ativos"), align="C",
                                  new_x="LMARGIN", new_y="NEXT")
                        self.set_font("Helvetica", "", 8)
                        self.cell(
                            0, 5,
                            _s(f"Emitido em: {datetime.date.today().strftime('%d/%m/%Y')}"
                               f"  |  Total: {_tot_pdf} aluno(s)"),
                            align="C", new_x="LMARGIN", new_y="NEXT",
                        )
                        _filtros_label = filtros if filtros else "Sem filtros"
                        self.set_font("Helvetica", "I", 7)
                        self.set_text_color(100, 116, 139)
                        self.cell(
                            0, 4,
                            _s(f"Filtros aplicados: {_filtros_label}"),
                            align="C", new_x="LMARGIN", new_y="NEXT",
                        )
                        self.set_text_color(15, 23, 42)
                        self.ln(2)
                    def footer(self):
                        self.set_y(-13)
                        self.set_font("Helvetica", "I", 7)
                        self.cell(0, 8, _s(f"Pag. {self.page_no()}"), align="C")

                pdf = _PDFHG(orientation="L", unit="mm", format="A4")
                pdf.add_page()
                pdf.set_auto_page_break(False, margin=14)

                # colunas dinâmicas conforme visibilidade
                _vp = lambda k: vis.get(k, True) if vis else True
                _pdf_col_spec = [
                    # (show, header, width)
                    (True,             "",               _FOTO_W),
                    (True,             "#",              6),
                    (True,             "Nome",           48),
                    (_vp("turma"),     "Turma",          20),
                    (_vp("aniv"),      "Aniversario",    13),
                    (_vp("freq"),      "Freq.Ano",       12),
                    (_vp("freq"),      "Ultima Pres.",   20),
                    (_vp("atestado"),  "Venc. Atestado", 20),
                    (_vp("pa"),        "Ult. PA",        28),
                    (_vp("anam"),      "Anamnese",       32),
                    (_vp("wap"),       "WhatsApp",       22),
                    (True,             "Voluntariado",   32),
                ]
                _hdrs   = [h for show, h, w in _pdf_col_spec if show]
                _widths = [w for show, h, w in _pdf_col_spec if show]

                # Cabeçalho azul
                pdf.set_font("Helvetica", "B", 8)
                pdf.set_fill_color(30, 77, 216)
                pdf.set_text_color(255, 255, 255)
                for _w, _h in zip(_widths, _hdrs):
                    pdf.cell(_w, 6, _s(_h), border=0, fill=True)
                pdf.ln()

                pdf.set_font("Helvetica", "", 8)
                pdf.set_text_color(15, 23, 42)
                _meses_pdf = ["","jan","fev","mar","abr","mai","jun",
                              "jul","ago","set","out","nov","dez"]
                _b_margin = 14

                for i, (_, a) in enumerate(df.iterrows()):
                    # Quebra de página manual (rect não dispara auto-break)
                    if pdf.get_y() + _ROW_H > pdf.h - _b_margin:
                        pdf.add_page()

                    _fr, _fg, _fb = (239, 246, 255) if i % 2 == 0 else (255, 255, 255)
                    pdf.set_fill_color(_fr, _fg, _fb)
                    pdf.set_text_color(15, 23, 42)

                    _y0 = pdf.get_y()
                    _x0 = pdf.l_margin

                    # Célula da foto: fundo + imagem circular
                    pdf.rect(_x0, _y0, _FOTO_W, _ROW_H, style="F")
                    _ibuf = _foto_cache.get(i)
                    if _ibuf:
                        try:
                            pdf.image(_ibuf,
                                      x=_x0 + _FOTO_PX,
                                      y=_y0 + _FOTO_PY,
                                      w=_FOTO_SZ, h=_FOTO_SZ)
                        except Exception:
                            pass

                    # Cursor após coluna foto; reaplica fill antes das células
                    pdf.set_xy(_x0 + _FOTO_W, _y0)
                    pdf.set_fill_color(_fr, _fg, _fb)

                    # Células de texto
                    _up = a.get("ultima_presenca")
                    try:
                        _up_str = pd.Timestamp(_up).strftime("%d/%m/%y") if pd.notna(_up) and _up else "-"
                    except Exception:
                        _up_str = "-"
                    _dv = a.get("data_vencimento_atestado")
                    try:
                        _dv_str = pd.Timestamp(_dv).strftime("%d/%m/%y") if pd.notna(_dv) and _dv else "-"
                    except Exception:
                        _dv_str = "-"
                    _dn = a.get("_dia_n"); _mn = a.get("_mes_n")
                    try:
                        _nasc_pdf = (f"{int(_dn):02d}/{_meses_pdf[int(_mn)]}"
                                     if pd.notna(_dn) and pd.notna(_mn) else "-")
                    except Exception:
                        _nasc_pdf = "-"
                    _anam_dt_pdf = a.get("_anam_data")
                    try:
                        _anam_pdf_str = (pd.Timestamp(_anam_dt_pdf).strftime("%d/%m/%y")
                                         if pd.notna(_anam_dt_pdf) else "-")
                    except Exception:
                        _anam_pdf_str = "-"
                    _anam_lbl_pdf = {
                        "nunca": "Nunca", "vencida": "Vencida",
                        "a_vencer": "A vencer", "ok": "Em dia",
                    }.get(str(a.get("_anam_status") or "nunca"), "-")
                    _vol_int_pdf   = str(a.get("trabalho_voluntario_interesse") or "").strip().lower()
                    _vol_acoes_pdf = _acoes_vol_dict.get(str(a.get("id", "")), [])
                    if _vol_acoes_pdf:
                        # Ações estruturadas — mostra ícone+nome abreviado
                        _vol_pdf_val = ", ".join(_vol_acoes_pdf)[:30]
                    elif _vol_int_pdf == "sim":
                        _fallback = str(a.get("trabalho_voluntario_areas") or "").strip()
                        _vol_pdf_val = _fallback[:22] if _fallback else "Sim"
                    else:
                        _vol_pdf_val = "-"
                    _txt_all = [
                        (True,             str(i + 1)),
                        (True,             str(a.get("nome", ""))[:28]),
                        (_vp("turma"),     str(a.get("turma", ""))[:12]),
                        (_vp("aniv"),      _nasc_pdf),
                        (_vp("freq"),      str(int(a.get("total_presencas_hist", 0)))),
                        (_vp("freq"),      _up_str),
                        (_vp("atestado"),  _dv_str),
                        (_vp("pa"),        str(a.get("_pa_txt", "") or "-")[:11]),
                        (_vp("anam"),      f"{_anam_pdf_str} ({_anam_lbl_pdf})"[:18]),
                        (_vp("wap"),       str(a.get("whatsapp") or "-")[:18]),
                        (True,             _vol_pdf_val),
                    ]
                    _txt_vals = [v for show, v in _txt_all if show]
                    for _pw, _pv in zip(_widths[1:], _txt_vals):
                        pdf.cell(_pw, _ROW_H, _s(_pv), border=0, fill=True)
                    pdf.ln()

                return bytes(pdf.output())

            # ── Gerador de Excel da lista home ───────────────────────────
            def _gerar_excel_hg(df: pd.DataFrame, vis: dict | None = None, filtros: str = "") -> bytes:
                import io as _io
                import openpyxl
                from openpyxl.styles import (
                    PatternFill, Font, Alignment, Border, Side, numbers as _opn
                )
                from openpyxl.utils import get_column_letter

                df = df.reset_index(drop=True)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Alunos Ativos"

                # ── Paleta ───────────────────────────────────────────────
                _HDR_BG   = "1E4DD8"   # azul cabeçalho
                _HDR_FG   = "FFFFFF"
                _EVEN_BG  = "EFF6FF"
                _ODD_BG   = "FFFFFF"
                _ALERT_BG = "FEF2F2"   # vermelho claro p/ atestado vencido
                _WARN_BG  = "FFFBEB"   # amarelo claro p/ a_vencer
                _OK_BG    = "F0FDF4"   # verde claro p/ ok

                def _fill(hex_color):
                    return PatternFill("solid", fgColor=hex_color)

                thin = Side(style="thin", color="D1D5DB")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                _meses = ["","jan","fev","mar","abr","mai","jun",
                          "jul","ago","set","out","nov","dez"]

                # Catálogo de colunas: (key, header, xl_width, color_key, wrap, center)
                _xl_spec = [
                    (None,       "#",               6,   None,    False, True),
                    (None,       "Nome",            36,  None,    True,  False),
                    ("turma",    "Turma",           14,  None,    False, False),
                    ("aniv",     "Aniversário",     13,  None,    False, False),
                    ("freq",     "Freq. Ano",        9,  None,    False, True),
                    ("freq",     "Total Presenças", 14,  None,    False, True),
                    ("freq",     "Última Presença", 14,  None,    False, False),
                    ("atestado", "Venc. Atestado",  14,  "atest", False, False),
                    ("atestado", "Status Atest.",   12,  "atest", False, False),
                    ("pa",       "Última PA",       12,  None,    False, False),
                    ("anam",     "Anamnese Data",   13,  "anam",  False, False),
                    ("anam",     "Status Anam.",    12,  "anam",  False, False),
                    (None,       "Voluntariado",    28,  None,    True,  False),
                    ("wap",      "WhatsApp",        16,  None,    True,  False),
                    (None,       "Tags de Saúde",   40,  None,    True,  False),
                ]
                _vxl = lambda k: k is None or (vis.get(k, True) if vis else True)
                _act_spec = [(k, h, w, ck, wr, cn) for k, h, w, ck, wr, cn in _xl_spec if _vxl(k)]
                _cols = [(h, w) for _, h, w, _, _, _ in _act_spec]

                # ── Título e metadados ────────────────────────────────────
                ws.merge_cells(f"A1:{get_column_letter(len(_cols))}1")
                ws["A1"] = f"IMBRA – Alunos Ativos  |  Emitido em: {datetime.date.today().strftime('%d/%m/%Y')}  |  Total: {len(df)} aluno(s)"
                ws["A1"].font = Font(bold=True, size=12, color=_HDR_FG)
                ws["A1"].fill = _fill(_HDR_BG)
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 22

                # ── Linha de filtros ativos ────────────────────────────────
                _filtros_label = filtros if filtros else "Sem filtros"
                ws.merge_cells(f"A2:{get_column_letter(len(_cols))}2")
                ws["A2"] = f"Filtros aplicados: {_filtros_label}"
                ws["A2"].font = Font(italic=True, size=9, color="475569")
                ws["A2"].fill = _fill("F1F5F9")
                ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[2].height = 16

                # ── Cabeçalhos ────────────────────────────────────────────
                _cols = [
                    ("#",              6),
                    ("Nome",           36),
                    ("Turma",          14),
                    ("Aniversário",    13),
                    ("Freq. Ano",        9),
                    ("Total Presenças", 14),
                    ("Última Presença", 14),
                    ("Venc. Atestado",  14),
                    ("Status Atest.",   12),
                    ("Última PA",       12),
                    ("Anamnese Data",   13),
                    ("Status Anam.",    12),
                    ("Voluntariado",    28),
                    ("WhatsApp",        16),
                    ("Tags de Saúde",   40),
                ]
                for ci, (lbl, w) in enumerate(_cols, start=1):
                    cell = ws.cell(row=3, column=ci, value=lbl)
                    cell.font      = Font(bold=True, color=_HDR_FG, size=9)
                    cell.fill      = _fill(_HDR_BG)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border    = border
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.row_dimensions[3].height = 26

                # Congelar título + filtros + cabeçalho
                ws.freeze_panes = "A4"

                # ── Linhas de dados ───────────────────────────────────────
                for i, (_, a) in enumerate(df.iterrows()):
                    row = i + 4
                    base_bg = _EVEN_BG if i % 2 == 0 else _ODD_BG

                    # --- campos calculados ---
                    _dn = a.get("_dia_n"); _mn = a.get("_mes_n")
                    try:
                        _aniv = (f"{int(_dn):02d}/{_meses[int(_mn)]}"
                                 if pd.notna(_dn) and pd.notna(_mn) else "")
                    except Exception:
                        _aniv = ""

                    _freq60  = int(a.get("total_presencas_hist", 0) or 0)
                    _tot_pr  = int(a.get("total_presencas_hist", 0) or 0)

                    _up = a.get("ultima_presenca")
                    try:
                        _up_str = pd.Timestamp(_up).strftime("%d/%m/%Y") if pd.notna(_up) and _up else ""
                    except Exception:
                        _up_str = ""

                    _dv = a.get("data_vencimento_atestado")
                    try:
                        _dv_ts = pd.Timestamp(_dv) if pd.notna(_dv) and _dv else None
                        _dv_str = _dv_ts.strftime("%d/%m/%Y") if _dv_ts else ""
                    except Exception:
                        _dv_ts = None
                        _dv_str = ""

                    # Calcular status do atestado a partir da data
                    try:
                        if _dv_ts:
                            _dv_dias = (_dv_ts.date() - datetime.date.today()).days
                            if _dv_dias < 0:
                                _atest_status = "vencido"
                            elif _dv_dias <= 30:
                                _atest_status = "a_vencer"
                            else:
                                _atest_status = "ok"
                        else:
                            _atest_status = ""
                    except Exception:
                        _atest_status = ""
                    _atest_lbl = {
                        "vencido": "Vencido ⚠️", "a_vencer": "A vencer",
                        "ok": "Em dia ✓", "": "-",
                    }.get(_atest_status, "-")

                    _anam_dt = a.get("_anam_data")
                    try:
                        _anam_str = pd.Timestamp(_anam_dt).strftime("%d/%m/%Y") if pd.notna(_anam_dt) else ""
                    except Exception:
                        _anam_str = ""
                    _anam_status = str(a.get("_anam_status") or "nunca")
                    _anam_lbl = {
                        "nunca": "Nunca", "vencida": "Vencida ⚠️",
                        "a_vencer": "A vencer", "ok": "Em dia ✓",
                    }.get(_anam_status, "-")

                    _pa_txt = str(a.get("_pa_txt") or "")

                    _acoes_vol = _acoes_vol_dict.get(str(a.get("id", "")), [])
                    if _acoes_vol:
                        _vol_str = ", ".join(_acoes_vol)
                    else:
                        _vol_int = str(a.get("trabalho_voluntario_interesse") or "").strip().lower()
                        if _vol_int == "sim":
                            _fb = str(a.get("trabalho_voluntario_areas") or "").strip()
                            _vol_str = _fb if _fb else "Sim"
                        else:
                            _vol_str = ""

                    _tags_str = str(a.get("tags_saude") or "")

                    _all_xl_vals = [
                        (None,       i + 1),
                        (None,       str(a.get("nome", ""))),
                        ("turma",    str(a.get("turma", ""))),
                        ("aniv",     _aniv),
                        ("freq",     _freq60),
                        ("freq",     _tot_pr),
                        ("freq",     _up_str),
                        ("atestado", _dv_str),
                        ("atestado", _atest_lbl),
                        ("pa",       _pa_txt),
                        ("anam",     _anam_str),
                        ("anam",     _anam_lbl),
                        (None,       _vol_str),
                        ("wap",      str(a.get("whatsapp") or "")),
                        (None,       _tags_str),
                    ]
                    vals = [v for k, v in _all_xl_vals if _vxl(k)]

                    _atest_cell_bg = {
                        "vencido": _ALERT_BG, "a_vencer": _WARN_BG, "ok": _OK_BG,
                    }.get(_atest_status, base_bg)
                    _anam_cell_bg = {
                        "vencida": _ALERT_BG, "a_vencer": _WARN_BG,
                        "ok": _OK_BG, "nunca": _ALERT_BG,
                    }.get(_anam_status, base_bg)

                    for ci, (val, (_, _, _, color_key, wrap, center)) in enumerate(
                        zip(vals, _act_spec), start=1
                    ):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.border = border
                        if color_key == "atest":
                            cell.fill = _fill(_atest_cell_bg)
                        elif color_key == "anam":
                            cell.fill = _fill(_anam_cell_bg)
                        else:
                            cell.fill = _fill(base_bg)
                        cell.alignment = Alignment(
                            horizontal="center" if center else "left",
                            vertical="center",
                            wrap_text=wrap,
                        )

                    ws.row_dimensions[row].height = 18

                # ── Auto-filtro ───────────────────────────────────────────
                ws.auto_filter.ref = f"A3:{get_column_letter(len(_cols))}3"

                buf = _io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            # Barra de info + navegação + impressão + excel
            _hg_pdf_key  = "hg_pdf_lista"
            _hg_xlsx_key = "hg_xlsx_lista"

            # ── Rótulo de filtros ativos para exportação ──────────────────
            _filtros_partes = []
            if _hg_turma and _hg_turma != "Todas":
                _filtros_partes.append(f"Turma: {_hg_turma}")
            if _hg_sexo and _hg_sexo != "Todos":
                _filtros_partes.append(_hg_sexo)
            if _hg_busca and len(_hg_busca.strip()) >= 3:
                _filtros_partes.append(f'Busca: "{_hg_busca.strip()}"')
            if _hg_faixa_ev:
                _faixa_desc_map = {
                    "verde":    "Ausência ≤ 7 dias",
                    "amarelo":  "Ausência 8–30 dias",
                    "laranja":  "Ausência 31–60 dias",
                    "vermelho": "Ausência > 60 dias",
                }
                _filtros_partes.append(_faixa_desc_map.get(_hg_faixa_ev, _hg_faixa_ev))
            if _hg_vol_filter:
                if _hg_acao_filter:
                    _filtros_partes.append(f"Voluntários: {', '.join(_hg_acao_filter)}")
                else:
                    _filtros_partes.append("Somente voluntários")
            _hg_filtros_label = " · ".join(_filtros_partes) if _filtros_partes else ""

            # ── Sufixo legível para o nome do arquivo exportado ───────────
            def _montar_sufixo_arquivo(partes):
                """Converte lista de partes de filtro em sufixo seguro para nome de arquivo."""
                if not partes:
                    return ""
                sufixo = "_".join(partes)
                # Substitui separadores e dois-pontos por underscore
                sufixo = sufixo.replace(": ", "_").replace(":", "_").replace(" · ", "_").replace(" - ", "_")
                # Remove ou substitui caracteres inválidos em nomes de arquivo
                sufixo = sufixo.replace(" ", "_")
                sufixo = re.sub(r'[\\/*?:"<>|≤≥–—""\'()]', "", sufixo)
                # Normaliza múltiplos underscores
                sufixo = re.sub(r"_+", "_", sufixo).strip("_")
                # Limita comprimento para não gerar nomes absurdamente longos
                return ("_" + sufixo[:60]) if sufixo else ""

            _hg_sufixo_arquivo = _montar_sufixo_arquivo(_filtros_partes)

            # Invalida cache PDF/Excel quando visibilidade ou filtros mudam
            _hg_vis_sig = str(sorted(_hg_vis.items())) + "|" + _hg_filtros_label
            if st.session_state.get("_hg_vis_sig_last") != _hg_vis_sig:
                st.session_state["_hg_vis_sig_last"] = _hg_vis_sig
                st.session_state.pop(_hg_pdf_key, None)
                st.session_state.pop(_hg_xlsx_key, None)
            _pc1, _pc2, _pc3, _pc_imp, _pc_xl = st.columns(
                [2.5, 0.6, 0.6, 0.9, 0.9], gap="small", vertical_alignment="center"
            )
            _pc1.caption(f"{len(_df_grid)} aluno(s) · pág {_hg_pg}/{_hg_total}")
            if _pc2.button("◀", key="hg_prev", disabled=(_hg_pg <= 1),
                           use_container_width=True):
                st.session_state.hg_pg -= 1
                st.rerun()
            if _pc3.button("▶", key="hg_next", disabled=(_hg_pg >= _hg_total),
                           use_container_width=True):
                st.session_state.hg_pg += 1
                st.rerun()

            # botão PDF
            if st.session_state.get(_hg_pdf_key):
                _pc_imp.download_button(
                    "📥 PDF",
                    data=st.session_state[_hg_pdf_key],
                    file_name=f"Alunos_Ativos_{datetime.date.today().strftime('%Y%m%d')}{_hg_sufixo_arquivo}.pdf",
                    mime="application/pdf",
                    key="hg_dl_pdf",
                    type="primary",
                    use_container_width=True,
                )
            else:
                if _pc_imp.button("🖨️ Imprimir", key="hg_imp", use_container_width=True):
                    st.session_state[_hg_pdf_key] = _gerar_pdf_hg(_df_grid, _hg_vis, _hg_filtros_label)
                    st.rerun()

            # botão Excel
            if st.session_state.get(_hg_xlsx_key):
                _pc_xl.download_button(
                    "📥 Excel",
                    data=st.session_state[_hg_xlsx_key],
                    file_name=f"Alunos_Ativos_{datetime.date.today().strftime('%Y%m%d')}{_hg_sufixo_arquivo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="hg_dl_xlsx",
                    type="secondary",
                    use_container_width=True,
                )
            else:
                if _pc_xl.button("📊 Excel", key="hg_xl", use_container_width=True):
                    st.session_state[_hg_xlsx_key] = _gerar_excel_hg(_df_grid, _hg_vis, _hg_filtros_label)
                    st.rerun()

            # ── Painel de evasão: contagem por faixa (clicável para filtrar) ────
            if _hg_vis.get("freq", True):
                # Conta faixas usando o dataframe COMPLETO filtrado por turma/busca/sexo
                # mas SEM o filtro de faixa, para mostrar totais reais por faixa
                _hoje_ev = datetime.date.today()
                _df_ev_base = _df_hg.copy()
                _df_ev_base = _faf_hg(_df_ev_base, _hg_busca, cols=["nome", "turma"], min_len=3)
                if _hg_turma != "Todas":
                    _df_ev_base = _df_ev_base[_df_ev_base["turma"] == _hg_turma]
                if _hg_sexo != "Todos" and "sexo" in _df_ev_base.columns:
                    _sexo_v2 = "F" if "Fem" in _hg_sexo else "M"
                    _df_ev_base = _df_ev_base[
                        _df_ev_base["sexo"].astype(str).str.upper().str.startswith(_sexo_v2)
                    ]
                _ev_verde = _ev_amarelo = _ev_laranja = _ev_vermelho = 0
                for _up_val in _df_ev_base.get("ultima_presenca", pd.Series(dtype="object")):
                    if pd.isna(_up_val) or _up_val is None:
                        _ev_vermelho += 1
                    else:
                        try:
                            _ev_d = (_hoje_ev - pd.Timestamp(_up_val).date()).days
                            if _ev_d <= 7:
                                _ev_verde += 1
                            elif _ev_d <= 30:
                                _ev_amarelo += 1
                            elif _ev_d <= 60:
                                _ev_laranja += 1
                            else:
                                _ev_vermelho += 1
                        except Exception:
                            _ev_vermelho += 1

                _faixa_ativa = st.session_state.get("hg_filtro_evasao")
                _ev_cols_parts = []
                if _ev_verde:   _ev_cols_parts.append(("verde",    f"🟢 {_ev_verde}",    "#D1FAE5", "#065F46"))
                if _ev_amarelo: _ev_cols_parts.append(("amarelo",  f"🟡 {_ev_amarelo}",  "#FEF3C7", "#92400E"))
                if _ev_laranja: _ev_cols_parts.append(("laranja",  f"🟠 {_ev_laranja}",  "#FFEDD5", "#9A3412"))
                if _ev_vermelho:_ev_cols_parts.append(("vermelho", f"🔴 {_ev_vermelho}", "#FEE2E2", "#991B1B"))

                if _ev_cols_parts:
                    _ev_n_cols = len(_ev_cols_parts) + (1 if _faixa_ativa else 0)
                    _ev_label_col, *_ev_btn_cols = st.columns(
                        [1.2] + [1] * _ev_n_cols, gap="small"
                    )
                    _ev_label_col.markdown(
                        "<span style='font-size:11px;color:#94A3B8;line-height:2;'>Ausência:</span>",
                        unsafe_allow_html=True,
                    )
                    for (_fk, _flabel, _fbg, _ffg), _ecol in zip(_ev_cols_parts, _ev_btn_cols):
                        _ativo = (_faixa_ativa == _fk)
                        if _ecol.button(
                            f"{'✓ ' if _ativo else ''}{_flabel}",
                            key=f"hg_ev_btn_{_fk}",
                            use_container_width=True,
                            help="Remover filtro de faixa" if _ativo else f"Filtrar grid por esta faixa de ausência",
                        ):
                            if _ativo:
                                st.session_state.pop("hg_filtro_evasao", None)
                            else:
                                st.session_state["hg_filtro_evasao"] = _fk
                            st.session_state.hg_pg = 1
                            st.rerun()
                    # Botão "Ver todos" quando há filtro ativo
                    if _faixa_ativa and _ev_n_cols > len(_ev_cols_parts):
                        if _ev_btn_cols[-1].button(
                            "✖ Ver todos",
                            key="hg_ev_btn_clear",
                            use_container_width=True,
                            help="Remover filtro de faixa de ausência",
                        ):
                            st.session_state.pop("hg_filtro_evasao", None)
                            st.session_state.hg_pg = 1
                            st.rerun()

            # ── Cabeçalho dinâmico das colunas (clicável para ordenar) ──────────
            _hdr_w = [0.5, 2.1]
            _hdr_k = ["_h_foto", "_h_nome"]
            for _hk, _, _hw in _HG_COL_DEFS:
                if _hg_vis.get(_hk, True):
                    _hdr_w.append(_hw)
                    _hdr_k.append(f"_hh_{_hk}")
            _hdr_w.append(0.9); _hdr_k.append("_h_btn")
            _hdr_cols_list = st.columns(_hdr_w, gap="small")
            _hcm = dict(zip(_hdr_k, _hdr_cols_list))
            _hcm["_h_foto"].markdown(" ", unsafe_allow_html=True)
            _hcm["_h_btn"].markdown(" ", unsafe_allow_html=True)
            if "_hh_wap" in _hcm:
                _hcm["_hh_wap"].markdown(
                    "<span style='font-size:12px;font-weight:600;color:#475569;'>📱 WhatsApp</span>",
                    unsafe_allow_html=True,
                )

            def _sort_btn(col_widget, label, sort_key):
                _ativo = (st.session_state.get("hg_sort_col") == sort_key)
                _seta  = (" ▲" if st.session_state.get("hg_sort_asc", True) else " ▼") if _ativo else ""
                if col_widget.button(
                    f"{label}{_seta}",
                    key=f"hdr_{sort_key}",
                    use_container_width=True,
                    help=f"Ordenar por {label}",
                ):
                    if st.session_state.get("hg_sort_col") == sort_key:
                        st.session_state.hg_sort_asc = not st.session_state.get("hg_sort_asc", True)
                    else:
                        st.session_state.hg_sort_col = sort_key
                        st.session_state.hg_sort_asc = True
                    st.session_state.hg_pg = 1
                    st.rerun()

            _sort_btn(_hcm["_h_nome"], "Nome", "nome")
            if "_hh_turma"    in _hcm: _sort_btn(_hcm["_hh_turma"],    "Turma",              "turma")
            if "_hh_aniv"     in _hcm: _sort_btn(_hcm["_hh_aniv"],     "🎂 Aniversário",     "aniversario")
            if "_hh_freq"     in _hcm:
                _hf1, _hf2 = _hcm["_hh_freq"].columns(2, gap="small")
                _sort_btn(_hf1, "⏱ Freq.Ano", "total_presencas_hist")
                _sort_btn(_hf2, "📅 Últ. Pres.", "ultima_presenca")
            if "_hh_atestado" in _hcm: _sort_btn(_hcm["_hh_atestado"], "🏥 Venc. Atestado",  "data_vencimento_atestado")
            if "_hh_pa"       in _hcm: _sort_btn(_hcm["_hh_pa"],       "🩸 Últ. PA",         "_pa_sis")
            if "_hh_anam"     in _hcm: _sort_btn(_hcm["_hh_anam"],     "📋 Anamnese",        "_anam_data")

            # Pré-carregar IDs avaliados (1 query, sem N+1 no loop)
            try:
                from database import get_ids_alunos_avaliados
                _ids_avaliados_hg = get_ids_alunos_avaliados()
            except Exception:
                _ids_avaliados_hg = set()

            # ── Linhas do Grid ─────────────────────────────────────────
            _hoje_hg   = datetime.date.today()
            _hoje_dia  = _hoje_hg.day
            _hoje_mes  = _hoje_hg.month
            _meses_abr = ["","jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]

            for _hg_row_idx, (_, _r) in enumerate(_df_pag.iterrows()):
                with st.container(border=True):
                    # ── Destaque visual por faixa de risco ─────────────────────
                    _up_hg_r = _r.get("ultima_presenca")
                    _row_faixa_hg = "verde"
                    if pd.isna(_up_hg_r) or _up_hg_r is None:
                        _row_faixa_hg = "vermelho"
                    else:
                        try:
                            _d_abs_hg = (_hoje_hg - pd.Timestamp(_up_hg_r).date()).days
                            if _d_abs_hg <= 7:
                                _row_faixa_hg = "verde"
                            elif _d_abs_hg <= 30:
                                _row_faixa_hg = "amarelo"
                            elif _d_abs_hg <= 60:
                                _row_faixa_hg = "laranja"
                            else:
                                _row_faixa_hg = "vermelho"
                        except Exception:
                            _row_faixa_hg = "vermelho"
                    _row_bg_hg = {"vermelho": "#FFF5F5", "laranja": "#FFF8F3"}.get(_row_faixa_hg, "")
                    if _row_bg_hg:
                        _row_css_cls = f"hgrow_{str(_r.get('id', _hg_row_idx)).replace('-', '_')}"
                        st.markdown(
                            f"<style>div[data-testid='stVerticalBlockBorderWrapper']"
                            f":has(div.{_row_css_cls}){{background:{_row_bg_hg}!important;}}</style>"
                            f"<div class='{_row_css_cls}' style='display:none'></div>",
                            unsafe_allow_html=True,
                        )

                    # Colunas dinâmicas conforme visibilidade
                    class _HgNoop:
                        def markdown(self, *a, **kw): pass
                    _HGNOOP = _HgNoop()
                    _row_w = [0.5, 2.1]
                    _row_k = ["_ca", "_cb"]
                    for _rk, _, _rw in _HG_COL_DEFS:
                        if _hg_vis.get(_rk, True):
                            _row_w.append(_rw)
                            _row_k.append(_rk)
                    _row_w.append(0.9); _row_k.append("_cf")
                    _row_cols_list = st.columns(_row_w, gap="small", vertical_alignment="center")
                    _rcm = dict(zip(_row_k, _row_cols_list))
                    _ca    = _rcm["_ca"]; _cb = _rcm["_cb"]; _cf = _rcm["_cf"]
                    _cc    = _rcm.get("turma",    _HGNOOP)
                    _cd    = _rcm.get("aniv",     _HGNOOP)
                    _ce    = _rcm.get("freq",     _HGNOOP)
                    _cg    = _rcm.get("atestado", _HGNOOP)
                    _cpa   = _rcm.get("pa",       _HGNOOP)
                    _canam = _rcm.get("anam",     _HGNOOP)
                    _cwap  = _rcm.get("wap",      _HGNOOP)
                    _caus  = _rcm.get("ausencia", _HGNOOP)

                    # Foto
                    _foto = str(_r.get("foto_url") or "").strip()
                    _inic = "".join(p[0].upper() for p in str(_r["nome"]).split()[:2] if p)
                    if _foto.startswith("http"):
                        _av_html = (
                            f'<div class="hg-avatar-wrap">'
                            f'<img class="hg-avatar" src="{_foto}" alt="{_inic}" '
                            f"onerror=\"this.outerHTML='<div class=hg-initials>{_inic}</div>'\">"
                            f'</div>'
                        )
                    else:
                        _av_html = f'<div class="hg-avatar-wrap"><div class="hg-initials">{_inic}</div></div>'
                    with _ca:
                        st.markdown(_av_html, unsafe_allow_html=True)

                    # Nome
                    _sem_img = not _r.get("termo_imagem")
                    _badge_img = (
                        "<abbr title='Não autoriza uso de imagem e voz (LGPD)' "
                        "style='text-decoration:none;cursor:help;"
                        "background:#FEE2E2;color:#DC2626;"
                        "border:1px solid #FECACA;"
                        "border-radius:3px;padding:1px 5px;"
                        "font-size:10px;font-weight:800;"
                        "margin-right:5px;white-space:nowrap;'>📷✕</abbr>"
                    ) if _sem_img else ""
                    _atestado_bloq = bool(_r.get("atestado_bloqueado"))
                    _badge_atestado = (
                        "<abbr title='Atestado médico pendente — Participação bloqueada' "
                        "style='text-decoration:none;cursor:help;"
                        "background:#FFF7ED;color:#C2410C;"
                        "border:1px solid #FED7AA;"
                        "border-radius:3px;padding:1px 5px;"
                        "font-size:10px;font-weight:800;"
                        "margin-right:5px;white-space:nowrap;'>🏥⚠</abbr>"
                    ) if _atestado_bloq else ""
                    _nunca_av = str(_r.get("id", "")) not in _ids_avaliados_hg
                    _badge_sem_av = (
                        "<abbr title='Nunca avaliado — sem nenhuma avaliação registrada' "
                        "style='text-decoration:none;cursor:help;"
                        "background:#EFF6FF;color:#1D4ED8;"
                        "border:1px solid #BFDBFE;"
                        "border-radius:3px;padding:1px 5px;"
                        "font-size:10px;font-weight:800;"
                        "margin-right:5px;white-space:nowrap;'>📋✕</abbr>"
                    ) if _nunca_av else ""
                    _aval_pend = bool(_r.get("avaliacao_pendente"))
                    _badge_aval_pend = (
                        "<abbr title='Bloqueado — aguardando reavaliação' "
                        "style='text-decoration:none;cursor:help;"
                        "background:#FEF3C7;color:#92400E;"
                        "border:1px solid #FCD34D;"
                        "border-radius:3px;padding:1px 5px;"
                        "font-size:10px;font-weight:800;"
                        "margin-right:5px;white-space:nowrap;'>⚡🚫</abbr>"
                    ) if _aval_pend else ""
                    # Badge voluntário
                    _vol_int_r   = str(_r.get("trabalho_voluntario_interesse") or "").strip().lower()
                    _vol_acoes_r = _acoes_vol_dict.get(str(_r.get("id", "")), [])
                    # Tooltip: ações estruturadas > texto livre > genérico
                    if _vol_acoes_r:
                        _vol_title = ", ".join(_vol_acoes_r)
                    else:
                        _fallback_areas = str(_r.get("trabalho_voluntario_areas") or "").strip()
                        _vol_title = _fallback_areas if _fallback_areas else "áreas não especificadas"
                    _badge_vol = (
                        f"<abbr title='Voluntária(o): {_vol_title}' "
                        "style='text-decoration:none;cursor:help;"
                        "background:#F0FDF4;color:#166534;"
                        "border:1px solid #86EFAC;"
                        "border-radius:3px;padding:1px 5px;"
                        "font-size:10px;font-weight:800;"
                        "margin-right:5px;white-space:nowrap;'>🤝 Vol.</abbr>"
                    ) if _vol_int_r == "sim" else ""
                    # Linha de ações (visível quando filtro de voluntários ativo)
                    if _vol_int_r == "sim" and _hg_vol_filter:
                        if _vol_acoes_r:
                            _pills = "".join(
                                f"<span style='display:inline-block;margin:1px 2px 0 0;"
                                f"font-size:10px;color:#166534;font-weight:600;"
                                f"background:#F0FDF4;border:1px solid #86EFAC;"
                                f"border-radius:10px;padding:1px 7px;'>{a}</span>"
                                for a in _vol_acoes_r
                            )
                            _vol_linha = f"<br>{_pills}"
                        elif str(_r.get("trabalho_voluntario_areas") or "").strip():
                            _vol_linha = (
                                f"<br><span style='font-size:10px;color:#166534;font-weight:600;"
                                f"background:#F0FDF4;border-radius:3px;padding:1px 5px;"
                                f"display:inline-block;margin-top:2px;'>"
                                f"🤝 {_r.get('trabalho_voluntario_areas','').strip()}</span>"
                            )
                        else:
                            _vol_linha = ""
                    else:
                        _vol_linha = ""
                    _cb.markdown(
                        f"<span style='font-size:13.5px;font-weight:700;color:#0F172A;'>"
                        f"{_badge_img}{_badge_atestado}{_badge_sem_av}{_badge_aval_pend}{_badge_vol}{_r['nome']}</span>"
                        f"{_vol_linha}",
                        unsafe_allow_html=True,
                    )

                    # Turma
                    _turma_v = str(_r.get("turma") or "—").strip() or "—"
                    _cc.markdown(
                        f"<span style='font-size:14px;font-weight:600;color:#475569;'>{_turma_v}</span>",
                        unsafe_allow_html=True,
                    )

                    # Aniversário
                    _dia_n = _r.get("_dia_n")
                    _mes_n = _r.get("_mes_n")
                    if pd.notna(_dia_n) and pd.notna(_mes_n):
                        _dia_n, _mes_n = int(_dia_n), int(_mes_n)
                        _nasc_str = f"{_dia_n:02d}/{_meses_abr[_mes_n]}"
                        if _mes_n == _hoje_mes and _dia_n == _hoje_dia:
                            _niver_html = f"<span class='hg-niver-hoje'>🎉 {_nasc_str} HOJE!</span>"
                        elif (_mes_n > _hoje_mes) or (_mes_n == _hoje_mes and _dia_n > _hoje_dia):
                            _niver_html = f"<span class='hg-niver-breve'>⏳ {_nasc_str}</span>"
                        else:
                            _niver_html = f"<span class='hg-niver-passou'>🎈 {_nasc_str}</span>"
                        if _wapp_ok and _mes_n == _hoje_mes:
                            _ja_parab_hg = str(_r.get("id", "")) in _parab_dict
                            if _ja_parab_hg:
                                _ts_hg = _parab_dict.get(str(_r.get("id", "")), "")
                                _ts_hg = _ts_hg.split("|")[0][:10].replace("T", " ") if _ts_hg else ""
                                _niver_html += (
                                    f"<br><span style='font-size:10px;color:#065F46;"
                                    f"font-weight:800;background:#D1FAE5;padding:1px 5px;"
                                    f"border-radius:4px;'>✅ Parabenizado</span>"
                                    + (f"<br><span style='font-size:9px;color:#94A3B8;'>{_ts_hg}</span>" if _ts_hg else "")
                                )
                            else:
                                _wap_n = str(_r.get("whatsapp") or "").strip()
                                if _mes_n == _hoje_mes and _dia_n == _hoje_dia:
                                    _niver_status = "hoje"
                                elif _dia_n < _hoje_dia:
                                    _niver_status = "passou"
                                else:
                                    _niver_status = None
                                if _wap_n and _niver_status:
                                    _msg_n = montar_mensagem_niver(_niver_status, str(_r.get("nome", "")))
                                    _link_n = montar_link_whatsapp(_wap_n, _msg_n)
                                    if _link_n:
                                        _niver_html += (
                                            f"<br><a href='{_link_n}' target='_blank' "
                                            f"style='font-size:10px;color:#25D366;text-decoration:none;"
                                            f"font-weight:600;'>📱 Parabéns</a>"
                                        )
                    else:
                        _niver_html = "<span style='color:#CBD5E1;'>—</span>"
                    _cd.markdown(
                        f"<span style='font-size:12px;'>{_niver_html}</span>",
                        unsafe_allow_html=True,
                    )

                    # Freq.60d + Última Presença (coluna combinada)
                    _tp_hist  = int(_r.get("total_presencas_hist", 0))
                    _tp_cor   = "#10B981" if _tp_hist >= 60 else ("#F59E0B" if _tp_hist >= 20 else "#EF4444")
                    _up = _r.get("ultima_presenca")
                    _dias_sem = ["seg","ter","qua","qui","sex","sáb","dom"]
                    if pd.notna(_up):
                        _up_dt   = pd.Timestamp(_up).date()
                        _up_dias = (_hoje_hg - _up_dt).days
                        if _up_dias <= 7:
                            _up_cor = "#10B981"; _up_badge_bg = "#D1FAE5"; _up_badge_fg = "#065F46"; _up_badge_ic = "🟢"
                        elif _up_dias <= 30:
                            _up_cor = "#F59E0B"; _up_badge_bg = "#FEF3C7"; _up_badge_fg = "#92400E"; _up_badge_ic = "🟡"
                        elif _up_dias <= 60:
                            _up_cor = "#F97316"; _up_badge_bg = "#FFEDD5"; _up_badge_fg = "#9A3412"; _up_badge_ic = "🟠"
                        else:
                            _up_cor = "#EF4444"; _up_badge_bg = "#FEE2E2"; _up_badge_fg = "#991B1B"; _up_badge_ic = "🔴"
                        _up_txt  = _up_dt.strftime("%d/%m/%y")
                        _up_dsem = _dias_sem[_up_dt.weekday()]
                        _up_html = (
                            f"<span style='font-size:10px;font-weight:700;background:{_up_badge_bg};"
                            f"color:{_up_badge_fg};border-radius:999px;padding:1px 6px;"
                            f"white-space:nowrap;'>{_up_badge_ic} {_up_dias}d</span> "
                            f"<span style='font-size:11px;font-weight:700;color:{_up_cor};'>"
                            f"{_up_txt}</span>"
                            f"<span style='font-size:10px;color:#94A3B8;margin-left:2px;'>{_up_dsem}</span>"
                        )
                        if _wapp_ok and _up_dias > 14:
                            _wap_e = str(_r.get("whatsapp") or "").strip()
                            if _wap_e:
                                _ev_key = "evasao_60" if _up_dias <= 30 else "evasao_80"
                                _tpl_e = _tpl_map.get(_ev_key, "Olá {nome}, sentimos sua falta!")
                                _msg_e = personalizar_mensagem(_tpl_e, str(_r.get("nome", "")))
                                _link_e = montar_link_whatsapp(_wap_e, _msg_e)
                                if _link_e:
                                    _cor_icone = "#F59E0B" if _up_dias <= 30 else "#EF4444"
                                    _up_html += (
                                        f"<br><a href='{_link_e}' target='_blank' "
                                        f"style='font-size:10px;color:{_cor_icone};text-decoration:none;"
                                        f"font-weight:600;'>📱 Contato</a>"
                                    )
                    else:
                        _up_html = (
                            "<span style='font-size:10px;font-weight:700;background:#FEE2E2;"
                            "color:#991B1B;border-radius:999px;padding:1px 6px;white-space:nowrap;'>"
                            "🔴 nunca</span>"
                        )
                        if _wapp_ok:
                            _wap_e = str(_r.get("whatsapp") or "").strip()
                            if _wap_e:
                                _tpl_e = _tpl_map.get("evasao_nunca", "Olá {nome}, ainda não te vimos por aqui!")
                                _msg_e = personalizar_mensagem(_tpl_e, str(_r.get("nome", "")))
                                _link_e = montar_link_whatsapp(_wap_e, _msg_e)
                                if _link_e:
                                    _up_html += (
                                        f"<br><a href='{_link_e}' target='_blank' "
                                        f"style='font-size:10px;color:#EF4444;text-decoration:none;"
                                        f"font-weight:600;'>📱 Contato</a>"
                                    )
                    _ce.markdown(
                        f"<div style='line-height:1.35;'>"
                        f"<span style='font-size:15px;font-weight:900;color:{_tp_cor};'>{_tp_hist}</span>"
                        f"<span style='font-size:9px;color:#94A3B8;margin-left:2px;'>aulas/ano</span>"
                        f"<br>{_up_html}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

                    # ⚠️ Ausência — badge de risco com link de acolhimento
                    _up_aus = _r.get("ultima_presenca")
                    if pd.notna(_up_aus):
                        _up_aus_dt   = pd.Timestamp(_up_aus).date()
                        _up_aus_dias = (_hoje_hg - _up_aus_dt).days
                        _aus_ic = (
                            "🟢" if _up_aus_dias <= 7 else
                            "🟡" if _up_aus_dias <= 29 else
                            "🟠" if _up_aus_dias <= 59 else "🔴"
                        )
                        _aus_bg = (
                            "#D1FAE5" if _up_aus_dias <= 7 else
                            "#FEF3C7" if _up_aus_dias <= 29 else
                            "#FFEDD5" if _up_aus_dias <= 59 else "#FEE2E2"
                        )
                        _aus_fg = (
                            "#065F46" if _up_aus_dias <= 7 else
                            "#92400E" if _up_aus_dias <= 29 else
                            "#9A3412" if _up_aus_dias <= 59 else "#991B1B"
                        )
                        _aus_html = (
                            f"<span style='font-size:11px;font-weight:700;"
                            f"background:{_aus_bg};color:{_aus_fg};"
                            f"border-radius:999px;padding:2px 8px;"
                            f"white-space:nowrap;'>{_aus_ic} {_up_aus_dias}d</span>"
                        )
                        if _wapp_ok and _up_aus_dias > 14:
                            _wap_aus = str(_r.get("whatsapp") or "").strip()
                            if _wap_aus:
                                _ev_key_aus = "evasao_60" if _up_aus_dias <= 30 else "evasao_80"
                                _tpl_aus = _tpl_map.get(
                                    _ev_key_aus, "Olá {nome}, sentimos sua falta!"
                                )
                                _msg_aus = personalizar_mensagem(
                                    _tpl_aus, str(_r.get("nome", ""))
                                )
                                _link_aus = montar_link_whatsapp(_wap_aus, _msg_aus)
                                if _link_aus:
                                    _cor_aus = "#F59E0B" if _up_aus_dias <= 30 else "#EF4444"
                                    _aus_html += (
                                        f"<br><a href='{_link_aus}' target='_blank' "
                                        f"style='font-size:10px;color:{_cor_aus};"
                                        f"text-decoration:none;font-weight:600;'>"
                                        f"📱 Acolher</a>"
                                    )
                    else:
                        _aus_html = (
                            "<span style='font-size:11px;font-weight:700;"
                            "background:#F1F5F9;color:#475569;"
                            "border-radius:999px;padding:2px 8px;white-space:nowrap;'>"
                            "⚫ nunca</span>"
                        )
                        if _wapp_ok:
                            _wap_aus = str(_r.get("whatsapp") or "").strip()
                            if _wap_aus:
                                _tpl_aus = _tpl_map.get(
                                    "evasao_nunca",
                                    "Olá {nome}, ainda não te vimos por aqui!"
                                )
                                _msg_aus = personalizar_mensagem(
                                    _tpl_aus, str(_r.get("nome", ""))
                                )
                                _link_aus = montar_link_whatsapp(_wap_aus, _msg_aus)
                                if _link_aus:
                                    _aus_html += (
                                        f"<br><a href='{_link_aus}' target='_blank' "
                                        f"style='font-size:10px;color:#EF4444;"
                                        f"text-decoration:none;font-weight:600;'>"
                                        f"📱 Acolher</a>"
                                    )
                    _caus.markdown(
                        f"<div style='line-height:1.5;'>{_aus_html}</div>",
                        unsafe_allow_html=True,
                    )

                    # Vencimento Atestado
                    _dv = _r.get("data_vencimento_atestado")
                    if pd.notna(_dv):
                        _dv_dt   = pd.Timestamp(_dv).date()
                        _dv_dias = (_dv_dt - _hoje_hg).days
                        if _dv_dias < 0:
                            _dv_cor   = "#DC2626"
                            _dv_bg    = "#FEE2E2"
                            _dv_icon  = "🔴"
                            _dv_label = f"Vencido {abs(_dv_dias)}d"
                        elif _dv_dias <= 30:
                            _dv_cor   = "#D97706"
                            _dv_bg    = "#FEF3C7"
                            _dv_icon  = "🟡"
                            _dv_label = f"{_dv_dias}d"
                        else:
                            _dv_cor   = "#059669"
                            _dv_bg    = "#D1FAE5"
                            _dv_icon  = "🟢"
                            _dv_label = f"{_dv_dias}d"
                        _dv_html = (
                            f"<span style='font-size:11px;font-weight:700;color:{_dv_cor};"
                            f"background:{_dv_bg};border-radius:4px;padding:2px 5px;"
                            f"display:inline-block;'>"
                            f"{_dv_icon} {_dv_dt.strftime('%d/%m/%y')}</span>"
                            f"<br><span style='font-size:10px;color:#94A3B8;'>{_dv_label}</span>"
                        )
                        if _wapp_ok and (_dv_dias < 0 or _dv_dias <= 30):
                            _wap_at = str(_r.get("whatsapp") or "").strip()
                            if _wap_at:
                                _at_key = "Atestado_Vencido" if _dv_dias < 0 else "Atestado_A_Vencer"
                                _at_default = (
                                    "Olá, {nome}! Seu atestado médico está vencido. Para continuar "
                                    "participando das atividades com segurança, pedimos que envie um "
                                    "atestado atualizado o quanto antes."
                                    if _dv_dias < 0 else
                                    "Olá, {nome}! Seu atestado médico vence em breve ({data_vencimento}). "
                                    "Para não haver interrupção nas suas atividades, pedimos que providencie "
                                    "a renovação."
                                )
                                _tpl_at = _tpl_map.get(_at_key, _at_default)
                                _msg_at = personalizar_mensagem(
                                    _tpl_at, str(_r.get("nome", "")),
                                    data_vencimento=_dv_dt.strftime("%d/%m/%Y"),
                                )
                                _link_at = montar_link_whatsapp(_wap_at, _msg_at)
                                if _link_at:
                                    _dv_html += (
                                        f"<br><a href='{_link_at}' target='_blank' "
                                        f"style='font-size:10px;color:#25D366;text-decoration:none;"
                                        f"font-weight:600;'>📱 Avisar</a>"
                                    )
                    else:
                        _dv_html = "<span style='font-size:11px;color:#CBD5E1;'>— sem atestado</span>"
                    _cg.markdown(_dv_html, unsafe_allow_html=True)

                    # Última PA compacta
                    _pa_html_v = str(_r.get("_pa_html", "<span style='color:#CBD5E1;font-size:11px;'>—</span>"))
                    _pa_cls_v  = str(_r.get("_pa_cls") or "")
                    if _wapp_ok and _pa_cls_v in ("estagio2", "crise", "estagio1"):
                        _wap_pa = str(_r.get("whatsapp") or "").strip()
                        if _wap_pa:
                            _pa_key = "PA_Grave" if _pa_cls_v in ("estagio2", "crise") else "PA_Atencao"
                            _pa_default = (
                                "Olá, {nome}! Identificamos que sua última aferição de pressão arterial "
                                "está em nível elevado ({pa_sistolica}/{pa_diastolica}). Recomendamos que "
                                "procure orientação médica o quanto antes."
                                if _pa_key == "PA_Grave" else
                                "Olá, {nome}! Sua última aferição de pressão arterial ficou um pouco acima "
                                "do ideal ({pa_sistolica}/{pa_diastolica}). Fique atento e, se possível, "
                                "converse com seu médico."
                            )
                            _tpl_pa = _tpl_map.get(_pa_key, _pa_default)
                            _msg_pa = personalizar_mensagem(
                                _tpl_pa, str(_r.get("nome", "")),
                                pa_sistolica=_r.get("_pa_sis"),
                                pa_diastolica=int(_r.get("_pa_dia")) if pd.notna(_r.get("_pa_dia")) else "",
                            )
                            _link_pa = montar_link_whatsapp(_wap_pa, _msg_pa)
                            if _link_pa:
                                _pa_html_v += (
                                    f"<br><a href='{_link_pa}' target='_blank' "
                                    f"style='font-size:10px;color:#25D366;text-decoration:none;"
                                    f"font-weight:600;'>📱 Avisar</a>"
                                )
                    _cpa.markdown(_pa_html_v, unsafe_allow_html=True)

                    # Status da Anamnese (avaliação clínica / prontuario_avaliacoes)
                    _anam_dt_v  = _r.get("_anam_data")
                    _anam_st_v  = str(_r.get("_anam_status") or "nunca")
                    if _anam_st_v == "nunca":
                        _anam_cor, _anam_bg, _anam_icon, _anam_label = "#DC2626", "#FEE2E2", "🔴", "Nunca feita"
                    elif _anam_st_v == "vencida":
                        _anam_cor, _anam_bg, _anam_icon, _anam_label = "#DC2626", "#FEE2E2", "🔴", "Vencida"
                    elif _anam_st_v == "a_vencer":
                        _anam_cor, _anam_bg, _anam_icon, _anam_label = "#D97706", "#FEF3C7", "🟡", "A vencer"
                    else:
                        _anam_cor, _anam_bg, _anam_icon, _anam_label = "#059669", "#D1FAE5", "🟢", "Em dia"
                    _anam_data_txt = (
                        pd.Timestamp(_anam_dt_v).strftime("%d/%m/%y") if pd.notna(_anam_dt_v) else "—"
                    )
                    _anam_html = (
                        f"<span style='font-size:11px;font-weight:700;color:{_anam_cor};"
                        f"background:{_anam_bg};border-radius:4px;padding:2px 5px;"
                        f"display:inline-block;'>{_anam_icon} {_anam_data_txt}</span>"
                        f"<br><span style='font-size:10px;color:#94A3B8;'>{_anam_label}</span>"
                    )
                    if _wapp_ok and _anam_st_v in ("nunca", "vencida", "a_vencer"):
                        _wap_an = str(_r.get("whatsapp") or "").strip()
                        if _wap_an:
                            _an_key = (
                                "Anamnese_Pendente" if _anam_st_v == "nunca"
                                else "Anamnese_Vencida" if _anam_st_v == "vencida"
                                else "Anamnese_A_Vencer"
                            )
                            _an_default = {
                                "Anamnese_Pendente": (
                                    "Olá, {nome}! Notamos que você ainda não realizou sua anamnese "
                                    "(avaliação clínica inicial). Para sua segurança, pedimos que agende "
                                    "esse atendimento o quanto antes."
                                ),
                                "Anamnese_Vencida": (
                                    "Olá, {nome}! Sua anamnese (avaliação clínica) está vencida. Para "
                                    "continuar participando das atividades com segurança, pedimos que "
                                    "agende uma reavaliação."
                                ),
                                "Anamnese_A_Vencer": (
                                    "Olá, {nome}! Sua anamnese (avaliação clínica) vencerá em breve. "
                                    "Pedimos que agende sua reavaliação para não haver interrupção nas "
                                    "suas atividades."
                                ),
                            }[_an_key]
                            _tpl_an = _tpl_map.get(_an_key, _an_default)
                            _msg_an = personalizar_mensagem(_tpl_an, str(_r.get("nome", "")))
                            _link_an = montar_link_whatsapp(_wap_an, _msg_an)
                            if _link_an:
                                _anam_html += (
                                    f"<br><a href='{_link_an}' target='_blank' "
                                    f"style='font-size:10px;color:#25D366;text-decoration:none;"
                                    f"font-weight:600;'>📱 Agendar</a>"
                                )
                    _canam.markdown(_anam_html, unsafe_allow_html=True)

                    # WhatsApp
                    _wap_v = str(_r.get("whatsapp") or "").strip()
                    if _wap_v:
                        _wap_digits = "".join(c for c in _wap_v if c.isdigit())
                        _wap_intl   = _wap_digits if _wap_digits.startswith("55") else f"55{_wap_digits}"
                        _wap_link   = f"https://wa.me/{_wap_intl}"
                        _wap_html   = (
                            f"<a href='{_wap_link}' target='_blank' "
                            f"style='font-size:12px;color:#25D366;font-weight:600;"
                            f"text-decoration:none;'>"
                            f"📱 {_wap_v}</a>"
                        )
                    else:
                        _wap_html = "<span style='color:#CBD5E1;font-size:12px;'>—</span>"
                    _cwap.markdown(_wap_html, unsafe_allow_html=True)

                    # Botão Ficha
                    with _cf:
                        if st.button(
                            "🩺", key=f"hg_{_r['id']}",
                            use_container_width=True, help="Abrir Prontuário"
                        ):
                            from database import buscar_aluno_por_id
                            _aluno_fresh = buscar_aluno_por_id(_r["id"]) or _r.to_dict()
                            st.session_state.aluno_prontuario = _aluno_fresh
                            st.session_state.origem_prontuario = "Principal"
                            st.session_state.menu_atual = "Portal do Aluno"
                            st.rerun()
        else:
            st.info("Nenhum aluno ativo encontrado.")

# ==============================================================================
# 🚀 ROTEAMENTO DE VISTAS
# ==============================================================================
elif st.session_state.menu_atual == "Frequência":
    _c_freq_titulo, _c_freq_btn = st.columns([6, 1], vertical_alignment="center")
    with _c_freq_btn:
        if _menu_liberado("freq_conf_facial") and st.button(
            "📸 Conf. Facial", use_container_width=True,
            help="Conferência de Presença por Foto"
        ):
            st.session_state.menu_atual = "Conferência Facial"
            st.rerun()
    from views.frequencia_view import tela_frequencia
    tela_frequencia()

elif st.session_state.menu_atual == "Nova Matrícula":
    st.session_state.menu_atual = "Portal do Aluno"
    st.rerun()

elif st.session_state.menu_atual == "Portal do Aluno":
    _col_portal, _col_ficha_portal = st.columns([6, 1], vertical_alignment="center")
    with _col_ficha_portal:
        if _menu_liberado("portal_ficha_impressao") and st.button(
            "🖨️ Ficha", use_container_width=True,
            help="Central de impressão de fichas de matrícula"
        ):
            st.session_state.menu_atual = "Ficha de Matrícula"
            st.rerun()
    if st.session_state.aluno_prontuario:
        if _menu_liberado("portal_prontuario"):
            from views.prontuario_ficha import renderizar_ficha
            renderizar_ficha()
        else:
            st.warning("🔒 Você não tem acesso ao prontuário individual. Contate o administrador.")
            st.session_state.aluno_prontuario = None
            st.rerun()
    else:
        from views.prontuario_dashboard import renderizar_dashboard
        renderizar_dashboard()

elif st.session_state.menu_atual in ("Relatórios & BI", "BI Prime", "Relatórios"):
    _col_rel, _col_ficha_rel = st.columns([6, 1], vertical_alignment="center")
    with _col_ficha_rel:
        if _menu_liberado("portal_ficha_impressao") and st.button(
            "🖨️ Ficha", use_container_width=True, key="ficha_btn_rel",
            help="Central de impressão de fichas de matrícula"
        ):
            st.session_state.menu_atual = "Ficha de Matrícula"
            st.rerun()
    # Monta apenas as abas que o operador tem acesso
    _rel_tabs_disponiveis = []
    if _menu_liberado("rel_relatorios"):
        _rel_tabs_disponiveis.append(("📋 Relatórios", "rel_relatorios"))
    if _menu_liberado("rel_bi_dashboard"):
        _rel_tabs_disponiveis.append(("📊 BI Dashboard", "rel_bi_dashboard"))
    if _menu_liberado("rel_bi_individual"):
        _rel_tabs_disponiveis.append(("👤 BI Individual", "rel_bi_individual"))
    if not _rel_tabs_disponiveis:
        st.info("🔒 Você não tem acesso a nenhuma aba de Relatórios & BI. Contate o administrador.")
    else:
        _aba_rel = st.tabs([_t[0] for _t in _rel_tabs_disponiveis])
        for _ri, (_rnome, _rchave) in enumerate(_rel_tabs_disponiveis):
            with _aba_rel[_ri]:
                if _rchave == "rel_relatorios":
                    from views.relatorio_view import tela_relatorio
                    tela_relatorio()
                elif _rchave == "rel_bi_dashboard":
                    from views.bi_dashboard_view import render_bi_dashboard
                    render_bi_dashboard()
                elif _rchave == "rel_bi_individual":
                    from views.bi_individual_view import render_bi_individual
                    render_bi_individual()

elif st.session_state.menu_atual == "Ficha de Matrícula":
    try:
        from views.ficha_aluno_view import tela_impressao_ficha

        tela_impressao_ficha()
    except:
        st.error("⚠️ Crie o ficheiro `ficha_aluno_view.py` na pasta `views`.")

elif st.session_state.menu_atual == "Conferência Facial":
    from views.conferencia_facial_view import tela_conferencia_facial
    tela_conferencia_facial()

elif st.session_state.menu_atual == "Radar de Inativos":
    from views.radar_acolhimento_view import tela_radar_acolhimento
    tela_radar_acolhimento()

elif st.session_state.menu_atual in (
    "Gestor",
    # ── Redirect de estados legados ─────────────────────────────────────
    "Radar de Acolhimento", "Satisfação",
    "Config", "Turmas", "Mensagens", "Identidade Visual", "Backup", "Mesclar Fichas",
):
    # ── Gestor: Radar · Satisfação · Emergência · Config (SuperAdmin) ──
    _e_super = st.session_state.get("perfil") == "SuperAdmin"
    _aba_g_nomes = []
    _aba_g_chaves = []
    for _nome_g, _chave_g in [
        ("💙 Radar", "gestor_radar"),
        ("⭐ Satisfação", "gestor_satisfacao"),
        ("🚨 Emergência", "gestor_emergencia"),
    ]:
        if _menu_liberado(_chave_g):
            _aba_g_nomes.append(_nome_g)
            _aba_g_chaves.append(_chave_g)
    if _e_super:
        _aba_g_nomes.append("⚙️ Config")
        _aba_g_chaves.append("_config")
    if not _aba_g_nomes:
        st.info("🔒 Você não tem acesso às abas do Gestor. Contate o administrador.")
        st.stop()
    _aba_g = st.tabs(_aba_g_nomes)

    def _idx_g(chave):
        """Retorna o índice da aba Gestor para a chave dada, ou None se não existir."""
        return _aba_g_chaves.index(chave) if chave in _aba_g_chaves else None

    if _idx_g("gestor_radar") is not None:
        with _aba_g[_idx_g("gestor_radar")]:
            from views.radar_acolhimento_view import tela_radar_acolhimento
            tela_radar_acolhimento()

    if _idx_g("gestor_satisfacao") is not None:
        with _aba_g[_idx_g("gestor_satisfacao")]:
            from views.relatorio_satisfacao_view import tela_relatorio_prime_satisfacao
            tela_relatorio_prime_satisfacao()

    if _idx_g("gestor_emergencia") is not None:
        with _aba_g[_idx_g("gestor_emergencia")]:
            from modulos_frequencia.tab_emergencia import renderizar_aba_emergencia
            renderizar_aba_emergencia(None, "Gestor")

    if _e_super:
        with _aba_g[_idx_g("_config")]:
            _aba_cfg = st.tabs([
                "🏫 Turmas", "💬 Mensagens", "🔔 Auto Niver", "🎨 Identidade Visual",
                "🗄️ Bck Adm", "🔀 Mescla Cad", "📅 Calendário", "📧 Email BI",
                "👥 Usuários", "🔒 LGPD", "🏷️ Tags Saúde", "🤝 Voluntariado",
                "🎊 Datas Comemorativas", "📡 Telemetria",
            ])
            with _aba_cfg[0]:
                from views.turmas_view import tela_gestao_turmas
                tela_gestao_turmas()
            with _aba_cfg[1]:
                from views.templates_view import tela_gestao_templates
                tela_gestao_templates()
            with _aba_cfg[2]:
                from views.config_niver_view import tela_config_niver
                tela_config_niver()
            with _aba_cfg[3]:
                from views.identidade_view import tela_identidade_visual
                tela_identidade_visual()
            with _aba_cfg[4]:
                from views.backup_view import tela_backup
                from database import ferramenta_reparacao_turmas
                tela_backup()
                st.markdown("---")
                ferramenta_reparacao_turmas()
            with _aba_cfg[5]:
                from views.merge_alunos_view import tela_merge_alunos
                tela_merge_alunos()
            with _aba_cfg[6]:
                _tela_calendario_institucional()
            with _aba_cfg[7]:
                _tela_email_bi()
            with _aba_cfg[8]:
                from views.gestao_usuarios_view import tela_gestao_usuarios
                tela_gestao_usuarios()
            with _aba_cfg[9]:
                from database import get_logs_lgpd
                st.markdown(
                    "<p style='font-weight:800;color:#0A2540;font-size:1rem;margin-bottom:4px;'>"
                    "🔒 Histórico de Alterações — Autorização de Imagem (LGPD)</p>",
                    unsafe_allow_html=True,
                )
                st.caption("Cada alteração do campo 'Uso de Imagem e Voz' registra automaticamente data/hora, operador e novo status.")
                _logs_lgpd = get_logs_lgpd()
                if not _logs_lgpd:
                    st.info("Nenhuma alteração registrada ainda.")
                else:
                    if st.button("🔄 Atualizar", key="lgpd_log_refresh"):
                        get_logs_lgpd.clear()
                        st.rerun()
                    for _lg in _logs_lgpd:
                        _ts_fmt = str(_lg.get("timestamp", ""))[:16].replace("T", " ")
                        _status_color = "#065F46" if _lg.get("status") == "Autorizado" else "#7F1D1D"
                        _status_bg = "#D1FAE5" if _lg.get("status") == "Autorizado" else "#FEE2E2"
                        _status_icon = "✅" if _lg.get("status") == "Autorizado" else "🚫"
                        st.markdown(
                            f"<div style='padding:8px 12px;border-radius:6px;margin-bottom:6px;"
                            f"background:#F8FAFC;border:1px solid #E2E8F0;display:flex;align-items:center;gap:12px;'>"
                            f"<span style='font-size:12px;color:#64748B;min-width:120px;'>🕒 {_ts_fmt}</span>"
                            f"<span style='font-size:13px;font-weight:700;color:#0F172A;flex:1;'>"
                            f"{_lg.get('aluno_nome', '—').upper()}</span>"
                            f"<span style='padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700;"
                            f"background:{_status_bg};color:{_status_color};'>"
                            f"{_status_icon} {_lg.get('status', '—')}</span>"
                            f"<span style='font-size:11px;color:#94A3B8;min-width:100px;text-align:right;'>"
                            f"op: {_lg.get('operador', '—')}</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

            with _aba_cfg[10]:
                from views.tags_clinicas_config_view import tela_tags_clinicas_config
                tela_tags_clinicas_config()

            with _aba_cfg[11]:
                from views.voluntariado_config_view import tela_voluntariado_config
                tela_voluntariado_config()

            with _aba_cfg[12]:
                from views.datas_comemorativas_view import tela_datas_comemorativas
                tela_datas_comemorativas()

            with _aba_cfg[13]:
                from views.telemetria_dashboard_view import tela_dashboard_telemetria
                tela_dashboard_telemetria()

# ── Rodapé Fixo ─────────────────────────────────────────────────────────────
from utils.identidade import get_config as _gcfg_rodape

_rcfg = _gcfg_rodape()
_rlinks = []
if _rcfg.get("site"):
    _rlinks.append(
        f'<a href="https://{_rcfg["site"]}" target="_blank">🌐 {_rcfg["site"]}</a>'
    )
if _rcfg.get("instagram"):
    _rlinks.append(
        f'<a href="https://instagram.com/{_rcfg["instagram"].lstrip("@")}"'
        f' target="_blank">📸 {_rcfg["instagram"]}</a>'
    )
st.markdown(
    f'<div class="rodape-prime">'
    f"<strong>{_rcfg.get('nome_organizacao', 'Instituto Muda Brasil')}</strong>"
    f"&nbsp;·&nbsp; MoveRight Elite"
    f"{'&nbsp;&nbsp;' + '&nbsp;|&nbsp;'.join(_rlinks) if _rlinks else ''}"
    f"</div>",
    unsafe_allow_html=True,
)
